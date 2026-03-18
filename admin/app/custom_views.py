"""Custom Russian admin views with consistent design"""
from flask import Blueprint, render_template_string, redirect, url_for, request, flash, Response
from flask_login import login_required, current_user
from app import db
from app.models import User, Level, UserLevelProgress, GameSession, UserActivity, AdminUser, GameText, LandingVisit, LandingStatsShare

bp = Blueprint('custom_admin', __name__)

# Shared base template
BASE_TEMPLATE = '''
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ title }} - ROSTIC'S Admin</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.0/font/bootstrap-icons.css" rel="stylesheet">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        :root {
            --primary: #6366f1;
            --primary-dark: #4f46e5;
            --success: #10b981;
            --warning: #f59e0b;
            --danger: #ef4444;
            --info: #06b6d4;
            --dark: #1e293b;
            --light: #f8fafc;
        }
        * { font-family: 'Inter', sans-serif; }
        body { background: var(--light); min-height: 100vh; }
        .sidebar {
            width: 260px;
            background: var(--dark);
            min-height: 100vh;
            position: fixed;
            left: 0;
            top: 0;
            padding: 1.5rem;
            z-index: 1000;
        }
        .sidebar-brand {
            color: white;
            font-size: 1.25rem;
            font-weight: 700;
            display: flex;
            align-items: center;
            gap: 0.75rem;
            padding-bottom: 1.5rem;
            border-bottom: 1px solid rgba(255,255,255,0.1);
            margin-bottom: 1.5rem;
            text-decoration: none;
        }
        .sidebar-brand:hover { color: white; }
        .sidebar-brand i { color: var(--primary); font-size: 1.5rem; }
        .nav-section {
            color: rgba(255,255,255,0.4);
            font-size: 0.7rem;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 1px;
            margin-bottom: 0.75rem;
            padding-left: 0.75rem;
        }
        .sidebar-nav { list-style: none; padding: 0; margin: 0 0 1.5rem 0; }
        .sidebar-nav a {
            color: rgba(255,255,255,0.7);
            text-decoration: none;
            display: flex;
            align-items: center;
            gap: 0.75rem;
            padding: 0.75rem;
            border-radius: 10px;
            transition: all 0.2s;
            font-weight: 500;
            font-size: 0.9rem;
        }
        .sidebar-nav a:hover { background: rgba(255,255,255,0.08); color: white; }
        .sidebar-nav a.active { background: var(--primary); color: white; }
        .sidebar-nav i { font-size: 1.1rem; opacity: 0.8; }
        .main-content { margin-left: 260px; padding: 2rem; }
        .page-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 2rem;
            flex-wrap: wrap;
            gap: 1rem;
        }
        .page-title { font-size: 1.75rem; font-weight: 700; color: var(--dark); margin: 0; }
        .card {
            background: white;
            border-radius: 16px;
            border: 1px solid #e2e8f0;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
            margin-bottom: 1.5rem;
        }
        .card-header {
            padding: 1.25rem 1.5rem;
            border-bottom: 1px solid #e2e8f0;
            display: flex;
            justify-content: space-between;
            align-items: center;
            background: transparent;
            flex-wrap: wrap;
            gap: 1rem;
        }
        .card-title { font-weight: 600; color: var(--dark); margin: 0; font-size: 1rem; }
        .card-body { padding: 1.5rem; }
        .table { margin: 0; }
        .table th {
            background: #f8fafc;
            font-weight: 600;
            font-size: 0.75rem;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            color: #64748b;
            padding: 1rem;
            border-bottom: 1px solid #e2e8f0;
            white-space: nowrap;
        }
        .table td {
            padding: 1rem;
            vertical-align: middle;
            border-bottom: 1px solid #f1f5f9;
            font-size: 0.9rem;
        }
        .table tbody tr:hover { background: #f8fafc; }
        .badge {
            font-weight: 500;
            padding: 0.35rem 0.75rem;
            border-radius: 6px;
            font-size: 0.75rem;
        }
        .badge-success { background: #dcfce7; color: #166534; }
        .badge-warning { background: #fef3c7; color: #92400e; }
        .badge-danger { background: #fee2e2; color: #991b1b; }
        .badge-info { background: #e0f2fe; color: #0369a1; }
        .btn {
            font-weight: 500;
            border-radius: 10px;
            padding: 0.5rem 1rem;
            font-size: 0.875rem;
            transition: all 0.2s;
        }
        .btn-primary { background: var(--primary); border: none; color: white; }
        .btn-primary:hover { background: var(--primary-dark); }
        .btn-danger { background: var(--danger); border: none; color: white; }
        .btn-success { background: var(--success); border: none; }
        .btn-outline-secondary { border: 2px solid #e2e8f0; color: var(--dark); background: white; }
        .btn-outline-secondary:hover { background: #f1f5f9; color: var(--dark); }
        .btn-sm { padding: 0.35rem 0.75rem; font-size: 0.8rem; }
        .form-control, .form-select {
            border: 2px solid #e2e8f0;
            border-radius: 10px;
            padding: 0.625rem 1rem;
            font-size: 0.9rem;
        }
        .form-control:focus, .form-select:focus {
            border-color: var(--primary);
            box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.15);
            outline: none;
        }
        .form-label { font-weight: 600; color: var(--dark); font-size: 0.875rem; }
        .pagination { margin: 0; }
        .page-link {
            border: none;
            color: var(--dark);
            padding: 0.5rem 0.875rem;
            border-radius: 8px;
            margin: 0 2px;
        }
        .page-link:hover { background: #f1f5f9; color: var(--dark); }
        .page-item.active .page-link { background: var(--primary); color: white; }
        .alert { border: none; border-radius: 12px; margin-bottom: 1.5rem; }
        .alert-success { background: #dcfce7; color: #166534; }
        .alert-danger { background: #fee2e2; color: #991b1b; }
        .search-box { max-width: 300px; }
        .action-btns { white-space: nowrap; }
        .action-btns .btn { margin-right: 0.25rem; }
        .table-responsive { border-radius: 0 0 16px 16px; }
        .stars { color: #f59e0b; }
        @media (max-width: 768px) {
            .sidebar { display: none; }
            .main-content { margin-left: 0; }
        }
    </style>
</head>
<body>
    <aside class="sidebar">
        <a href="/admin" class="sidebar-brand">
            <i class="bi bi-controller"></i>
            <span>ROSTIC'S Admin</span>
        </a>
        <div class="nav-section">Главная</div>
        <ul class="sidebar-nav">
            <li><a href="/admin" {% if active_page == 'dashboard' %}class="active"{% endif %}><i class="bi bi-grid-1x2"></i> Дашборд</a></li>
            <li><a href="/admin/analytics/" {% if active_page == 'analytics' %}class="active"{% endif %}><i class="bi bi-bar-chart-line"></i> Аналитика</a></li>
            <li><a href="/admin/landing-stats/" {% if active_page == 'landing_stats' %}class="active"{% endif %}><i class="bi bi-flower2"></i> Sakura Fest</a></li>
        </ul>
        {% if admin_role in ('superadmin', 'game_admin', 'quest_admin') %}
        <div class="nav-section">Пользователи</div>
        <ul class="sidebar-nav">
            <li><a href="/admin/user/" {% if active_page == 'users' %}class="active"{% endif %}><i class="bi bi-people"></i> Все пользователи</a></li>
            <li><a href="/admin/progress/" {% if active_page == 'progress' %}class="active"{% endif %}><i class="bi bi-graph-up"></i> Прогресс</a></li>
            <li><a href="/admin/activity/" {% if active_page == 'activity' %}class="active"{% endif %}><i class="bi bi-activity"></i> Активность</a></li>
        </ul>
        {% endif %}
        {% if admin_role in ('superadmin', 'game_admin') %}
        <div class="nav-section">Игра</div>
        <ul class="sidebar-nav">
            <li><a href="/admin/level/" {% if active_page == 'levels' %}class="active"{% endif %}><i class="bi bi-layers"></i> Уровни</a></li>
            <li><a href="/level-editor/" {% if active_page == 'level_editor' %}class="active"{% endif %}><i class="bi bi-grid-3x3-gap"></i> Конструктор</a></li>
            <li><a href="/admin/session/" {% if active_page == 'sessions' %}class="active"{% endif %}><i class="bi bi-joystick"></i> Сессии</a></li>
            <li><a href="/admin/texts/" {% if active_page == 'texts' %}class="active"{% endif %}><i class="bi bi-fonts"></i> Тексты</a></li>
        </ul>
        {% endif %}
        {% if admin_role in ('superadmin', 'quest_admin') %}
        <div class="nav-section">Квест</div>
        <ul class="sidebar-nav">
            <li><a href="/admin/quest/pages/" {% if active_page == 'quest_pages' %}class="active"{% endif %}><i class="bi bi-map"></i> Страницы квеста</a></li>
            <li><a href="/admin/quest/promo/" {% if active_page == 'quest_promo' %}class="active"{% endif %}><i class="bi bi-ticket-perforated"></i> Промокоды</a></li>
            <li><a href="/admin/quest/progress/" {% if active_page == 'quest_progress' %}class="active"{% endif %}><i class="bi bi-person-check"></i> Прогресс участников</a></li>
            <li><a href="/admin/quest/qr-codes/" {% if active_page == 'quest_qr' %}class="active"{% endif %}><i class="bi bi-qr-code"></i> QR-коды</a></li>
        </ul>
        {% endif %}
        <div class="nav-section">Система</div>
        <ul class="sidebar-nav">
            <li><a href="/admin/adminuser/" {% if active_page == 'admins' %}class="active"{% endif %}><i class="bi bi-shield-lock"></i> Админы</a></li>
            <li><a href="/auth/logout"><i class="bi bi-box-arrow-left"></i> Выход</a></li>
        </ul>
    </aside>
    <main class="main-content">
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, message in messages %}
                    <div class="alert alert-{{ 'success' if category == 'success' else 'danger' }}">
                        <i class="bi bi-{{ 'check-circle' if category == 'success' else 'exclamation-circle' }} me-2"></i>{{ message }}
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}
        {{ content | safe }}
    </main>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
'''


def render_page(title, active_page, content):
    from flask_login import current_user
    role = current_user.role if hasattr(current_user, 'role') and current_user.role else 'superadmin'
    return render_template_string(BASE_TEMPLATE, title=title, active_page=active_page, content=content, admin_role=role)


# ============== USERS ==============
@bp.route('/user/')
@login_required
def users_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    city_filter = request.args.get('city', '')

    query = User.query
    if search:
        query = query.filter(
            (User.username.ilike(f'%{search}%')) |
            (User.email.ilike(f'%{search}%'))
        )
    if city_filter:
        query = query.filter(User.city == city_filter)

    pagination = query.order_by(User.created_at.desc()).paginate(page=page, per_page=25, error_out=False)

    # Count users by city
    moscow_count = User.query.filter(User.city == 'moscow').count()
    region_count = User.query.filter(User.city == 'region').count()

    content = f'''
    <div class="page-header">
        <h1 class="page-title"><i class="bi bi-people me-2"></i>Пользователи</h1>
    </div>

    <!-- City Stats Cards -->
    <div class="row mb-4">
        <div class="col-md-4 col-sm-6 mb-3">
            <a href="/admin/user/" class="text-decoration-none">
                <div class="card h-100" style="border-left: 4px solid var(--primary); {'background: rgba(99, 102, 241, 0.05);' if not city_filter else ''}">
                    <div class="card-body py-3">
                        <div class="d-flex justify-content-between align-items-center">
                            <div>
                                <div class="text-muted small">Все</div>
                                <div style="font-size: 1.5rem; font-weight: 700; color: var(--dark);">{pagination.total if not city_filter else moscow_count + region_count:,}</div>
                            </div>
                            <i class="bi bi-people" style="font-size: 1.5rem; color: var(--primary);"></i>
                        </div>
                    </div>
                </div>
            </a>
        </div>
        <div class="col-md-4 col-sm-6 mb-3">
            <a href="/admin/user/?city=moscow" class="text-decoration-none">
                <div class="card h-100" style="border-left: 4px solid var(--danger); {'background: rgba(239, 68, 68, 0.05);' if city_filter == 'moscow' else ''}">
                    <div class="card-body py-3">
                        <div class="d-flex justify-content-between align-items-center">
                            <div>
                                <div class="text-muted small">🏙️ Москва и МО</div>
                                <div style="font-size: 1.5rem; font-weight: 700; color: var(--dark);">{moscow_count:,}</div>
                            </div>
                            <i class="bi bi-buildings" style="font-size: 1.5rem; color: var(--danger);"></i>
                        </div>
                    </div>
                </div>
            </a>
        </div>
        <div class="col-md-4 col-sm-6 mb-3">
            <a href="/admin/user/?city=region" class="text-decoration-none">
                <div class="card h-100" style="border-left: 4px solid var(--success); {'background: rgba(16, 185, 129, 0.05);' if city_filter == 'region' else ''}">
                    <div class="card-body py-3">
                        <div class="d-flex justify-content-between align-items-center">
                            <div>
                                <div class="text-muted small">🌍 Регионы</div>
                                <div style="font-size: 1.5rem; font-weight: 700; color: var(--dark);">{region_count:,}</div>
                            </div>
                            <i class="bi bi-geo-alt" style="font-size: 1.5rem; color: var(--success);"></i>
                        </div>
                    </div>
                </div>
            </a>
        </div>
    </div>

    <div class="card">
        <div class="card-header">
            <form method="GET" class="d-flex gap-2 flex-wrap">
                <input type="text" name="search" class="form-control" style="max-width: 200px;" placeholder="Поиск..." value="{search}">
                <input type="hidden" name="city" value="{city_filter}">
                <button type="submit" class="btn btn-primary"><i class="bi bi-search"></i></button>
            </form>
            <span class="text-muted">
                {f'🏙️ Москва и МО: ' if city_filter == 'moscow' else f'🌍 Регионы: ' if city_filter == 'region' else 'Всего: '}
                {pagination.total}
            </span>
        </div>
        <div class="table-responsive">
            <table class="table">
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Имя</th>
                        <th>Email</th>
                        <th>Город</th>
                        <th>Статус</th>
                        <th>Очки</th>
                        <th>Регистрация</th>
                        <th>Действия</th>
                    </tr>
                </thead>
                <tbody>
    '''

    for u in pagination.items:
        status = '<span class="badge badge-success">Подтверждён</span>' if u.is_verified else '<span class="badge badge-warning">Ожидает</span>'
        city_display = u.city_name or ('Москва и МО' if u.city == 'moscow' else 'Регион')
        city_badge_class = 'badge-danger' if u.city == 'moscow' else 'badge-success'
        date = u.created_at.strftime('%d.%m.%Y %H:%M') if u.created_at else '—'
        content += f'''
                    <tr>
                        <td><strong>#{u.id}</strong></td>
                        <td>{u.username}</td>
                        <td>{u.email}</td>
                        <td><span class="badge {city_badge_class}">{city_display}</span></td>
                        <td>{status}</td>
                        <td><strong>{u.total_score:,}</strong></td>
                        <td>{date}</td>
                        <td class="action-btns">
                            <a href="/admin/user/edit/{u.id}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
                            <a href="/admin/user/delete/{u.id}" class="btn btn-sm btn-danger" onclick="return confirm('Удалить пользователя?')"><i class="bi bi-trash"></i></a>
                        </td>
                    </tr>
        '''

    content += '''
                </tbody>
            </table>
        </div>
    '''

    if pagination.pages > 1:
        content += '<div class="card-footer bg-transparent border-0 d-flex justify-content-center"><nav><ul class="pagination">'
        if pagination.has_prev:
            content += f'<li class="page-item"><a class="page-link" href="?page={pagination.prev_num}&search={search}&city={city_filter}"><i class="bi bi-chevron-left"></i></a></li>'
        for p in pagination.iter_pages():
            if p:
                active = 'active' if p == pagination.page else ''
                content += f'<li class="page-item {active}"><a class="page-link" href="?page={p}&search={search}&city={city_filter}">{p}</a></li>'
            else:
                content += '<li class="page-item disabled"><span class="page-link">...</span></li>'
        if pagination.has_next:
            content += f'<li class="page-item"><a class="page-link" href="?page={pagination.next_num}&search={search}&city={city_filter}"><i class="bi bi-chevron-right"></i></a></li>'
        content += '</ul></nav></div>'

    content += '</div>'

    return render_page('Пользователи', 'users', content)


@bp.route('/user/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def user_edit(id):
    user = User.query.get_or_404(id)

    if request.method == 'POST':
        user.username = request.form.get('username', user.username)
        user.email = request.form.get('email', user.email)
        user.city = request.form.get('city', user.city)
        user.is_verified = request.form.get('is_verified') == '1'
        user.total_score = int(request.form.get('total_score', 0))
        user.verification_code = request.form.get('verification_code') or None
        db.session.commit()
        flash('Пользователь обновлён!', 'success')
        return redirect(url_for('custom_admin.users_list'))

    content = f'''
    <div class="page-header">
        <h1 class="page-title"><i class="bi bi-pencil me-2"></i>Редактирование пользователя #{user.id}</h1>
    </div>
    <div class="card">
        <div class="card-body">
            <form method="POST">
                <div class="row">
                    <div class="col-md-6 mb-3">
                        <label class="form-label">Имя</label>
                        <input type="text" name="username" class="form-control" value="{user.username}" required>
                    </div>
                    <div class="col-md-6 mb-3">
                        <label class="form-label">Email</label>
                        <input type="email" name="email" class="form-control" value="{user.email}" required>
                    </div>
                    <div class="col-md-4 mb-3">
                        <label class="form-label">Город</label>
                        <select name="city" class="form-select">
                            <option value="moscow" {'selected' if user.city == 'moscow' else ''}>🏙️ Москва и МО</option>
                            <option value="region" {'selected' if user.city == 'region' else ''}>🌍 Регионы</option>
                        </select>
                    </div>
                    <div class="col-md-4 mb-3">
                        <label class="form-label">Статус</label>
                        <select name="is_verified" class="form-select">
                            <option value="1" {'selected' if user.is_verified else ''}>Подтверждён</option>
                            <option value="0" {'' if user.is_verified else 'selected'}>Ожидает</option>
                        </select>
                    </div>
                    <div class="col-md-4 mb-3">
                        <label class="form-label">Очки</label>
                        <input type="number" name="total_score" class="form-control" value="{user.total_score}">
                    </div>
                    <div class="col-md-4 mb-3">
                        <label class="form-label">Код верификации</label>
                        <input type="text" name="verification_code" class="form-control" value="{user.verification_code or ''}">
                    </div>
                </div>
                <div class="d-flex gap-2">
                    <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-2"></i>Сохранить</button>
                    <a href="/admin/user/" class="btn btn-outline-secondary">Отмена</a>
                </div>
            </form>
        </div>
    </div>
    '''
    return render_page('Редактирование пользователя', 'users', content)


@bp.route('/user/delete/<int:id>')
@login_required
def user_delete(id):
    user = User.query.get_or_404(id)
    db.session.delete(user)
    db.session.commit()
    flash('Пользователь удалён!', 'success')
    return redirect(url_for('custom_admin.users_list'))


# ============== PROGRESS ==============
@bp.route('/progress/')
@login_required
def progress_list():
    page = request.args.get('page', 1, type=int)
    pagination = UserLevelProgress.query.order_by(UserLevelProgress.completed_at.desc().nullslast()).paginate(page=page, per_page=25, error_out=False)

    content = '''
    <div class="page-header">
        <h1 class="page-title"><i class="bi bi-graph-up me-2"></i>Прогресс игроков</h1>
    </div>
    <div class="card">
        <div class="card-header">
            <span class="card-title">История прохождений</span>
            <span class="text-muted">Всего: ''' + str(pagination.total) + '''</span>
        </div>
        <div class="table-responsive">
            <table class="table">
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Игрок</th>
                        <th>Уровень</th>
                        <th>Лучший счёт</th>
                        <th>Звёзды</th>
                        <th>Попытки</th>
                        <th>Завершён</th>
                    </tr>
                </thead>
                <tbody>
    '''

    for p in pagination.items:
        stars = '⭐' * p.stars + '☆' * (3 - p.stars) if p.stars else '—'
        date = p.completed_at.strftime('%d.%m.%Y %H:%M') if p.completed_at else '—'
        user_name = p.user.username if p.user else '—'
        level_name = p.level.name if p.level else '—'
        content += f'''
                    <tr>
                        <td><strong>#{p.id}</strong></td>
                        <td>{user_name}</td>
                        <td>{level_name}</td>
                        <td><strong>{p.best_score:,}</strong></td>
                        <td class="stars">{stars}</td>
                        <td>{p.attempts_count}</td>
                        <td>{date}</td>
                    </tr>
        '''

    content += '</tbody></table></div>'

    if pagination.pages > 1:
        content += '<div class="card-footer bg-transparent border-0 d-flex justify-content-center"><nav><ul class="pagination">'
        if pagination.has_prev:
            content += f'<li class="page-item"><a class="page-link" href="?page={pagination.prev_num}"><i class="bi bi-chevron-left"></i></a></li>'
        for pg in pagination.iter_pages():
            if pg:
                active = 'active' if pg == pagination.page else ''
                content += f'<li class="page-item {active}"><a class="page-link" href="?page={pg}">{pg}</a></li>'
        if pagination.has_next:
            content += f'<li class="page-item"><a class="page-link" href="?page={pagination.next_num}"><i class="bi bi-chevron-right"></i></a></li>'
        content += '</ul></nav></div>'

    content += '</div>'
    return render_page('Прогресс', 'progress', content)


# ============== ACTIVITY ==============
@bp.route('/activity/')
@login_required
def activity_list():
    page = request.args.get('page', 1, type=int)
    pagination = UserActivity.query.order_by(UserActivity.created_at.desc()).paginate(page=page, per_page=25, error_out=False)

    content = '''
    <div class="page-header">
        <h1 class="page-title"><i class="bi bi-activity me-2"></i>Активность</h1>
    </div>
    <div class="card">
        <div class="card-header">
            <span class="card-title">Журнал действий</span>
            <span class="text-muted">Всего: ''' + str(pagination.total) + '''</span>
        </div>
        <div class="table-responsive">
            <table class="table">
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Пользователь</th>
                        <th>Действие</th>
                        <th>IP</th>
                        <th>Дата</th>
                    </tr>
                </thead>
                <tbody>
    '''

    for a in pagination.items:
        date = a.created_at.strftime('%d.%m.%Y %H:%M') if a.created_at else '—'
        user_name = a.user.username if a.user else '—'
        content += f'''
                    <tr>
                        <td><strong>#{a.id}</strong></td>
                        <td>{user_name}</td>
                        <td><span class="badge badge-info">{a.action}</span></td>
                        <td><code>{a.ip_address or '—'}</code></td>
                        <td>{date}</td>
                    </tr>
        '''

    content += '</tbody></table></div>'

    if pagination.pages > 1:
        content += '<div class="card-footer bg-transparent border-0 d-flex justify-content-center"><nav><ul class="pagination">'
        if pagination.has_prev:
            content += f'<li class="page-item"><a class="page-link" href="?page={pagination.prev_num}"><i class="bi bi-chevron-left"></i></a></li>'
        for pg in pagination.iter_pages():
            if pg:
                active = 'active' if pg == pagination.page else ''
                content += f'<li class="page-item {active}"><a class="page-link" href="?page={pg}">{pg}</a></li>'
        if pagination.has_next:
            content += f'<li class="page-item"><a class="page-link" href="?page={pagination.next_num}"><i class="bi bi-chevron-right"></i></a></li>'
        content += '</ul></nav></div>'

    content += '</div>'
    return render_page('Активность', 'activity', content)


# ============== LEVELS ==============
@bp.route('/level/')
@login_required
def levels_list():
    levels = Level.query.order_by(Level.order).all()

    content = '''
    <div class="page-header">
        <h1 class="page-title"><i class="bi bi-layers me-2"></i>Уровни</h1>
        <a href="/level-editor/" class="btn btn-primary"><i class="bi bi-plus-lg me-2"></i>Создать уровень</a>
    </div>
    <div class="card">
        <div class="card-header">
            <span class="card-title">Все уровни</span>
            <span class="text-muted">Всего: ''' + str(len(levels)) + '''</span>
        </div>
        <div class="table-responsive">
            <table class="table">
                <thead>
                    <tr>
                        <th>№</th>
                        <th>Название</th>
                        <th>Размер</th>
                        <th>Ходы</th>
                        <th>Предметы</th>
                        <th>Статус</th>
                        <th>Действия</th>
                    </tr>
                </thead>
                <tbody>
    '''

    for lvl in levels:
        status = '<span class="badge badge-success">Активен</span>' if lvl.is_active else '<span class="badge badge-danger">Отключён</span>'
        items = ', '.join(lvl.item_types) if lvl.item_types else '—'
        content += f'''
                    <tr>
                        <td><strong>{lvl.order}</strong></td>
                        <td>{lvl.name}</td>
                        <td>{lvl.grid_width}×{lvl.grid_height}</td>
                        <td>{lvl.max_moves}</td>
                        <td><small>{items}</small></td>
                        <td>{status}</td>
                        <td class="action-btns">
                            <a href="/level-editor/{lvl.id}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
                        </td>
                    </tr>
        '''

    content += '</tbody></table></div></div>'
    return render_page('Уровни', 'levels', content)


# ============== SESSIONS ==============
@bp.route('/session/')
@login_required
def sessions_list():
    page = request.args.get('page', 1, type=int)
    pagination = GameSession.query.order_by(GameSession.created_at.desc()).paginate(page=page, per_page=25, error_out=False)

    content = '''
    <div class="page-header">
        <h1 class="page-title"><i class="bi bi-joystick me-2"></i>Игровые сессии</h1>
    </div>
    <div class="card">
        <div class="card-header">
            <span class="card-title">История игр</span>
            <span class="text-muted">Всего: ''' + str(pagination.total) + '''</span>
        </div>
        <div class="table-responsive">
            <table class="table">
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Игрок</th>
                        <th>Уровень</th>
                        <th>Очки</th>
                        <th>Ходов</th>
                        <th>Результат</th>
                        <th>Время</th>
                        <th>Дата</th>
                    </tr>
                </thead>
                <tbody>
    '''

    for s in pagination.items:
        result = '<span class="badge badge-success">Победа</span>' if s.is_won else '<span class="badge badge-danger">Проигрыш</span>'
        date = s.created_at.strftime('%d.%m.%Y %H:%M') if s.created_at else '—'
        user_name = s.user.username if s.user else '—'
        level_name = s.level.name if s.level else '—'
        content += f'''
                    <tr>
                        <td><strong>#{s.id}</strong></td>
                        <td>{user_name}</td>
                        <td>{level_name}</td>
                        <td><strong>{s.score:,}</strong></td>
                        <td>{s.moves_used}</td>
                        <td>{result}</td>
                        <td>{s.duration_seconds} сек</td>
                        <td>{date}</td>
                    </tr>
        '''

    content += '</tbody></table></div>'

    if pagination.pages > 1:
        content += '<div class="card-footer bg-transparent border-0 d-flex justify-content-center"><nav><ul class="pagination">'
        if pagination.has_prev:
            content += f'<li class="page-item"><a class="page-link" href="?page={pagination.prev_num}"><i class="bi bi-chevron-left"></i></a></li>'
        for pg in pagination.iter_pages():
            if pg:
                active = 'active' if pg == pagination.page else ''
                content += f'<li class="page-item {active}"><a class="page-link" href="?page={pg}">{pg}</a></li>'
        if pagination.has_next:
            content += f'<li class="page-item"><a class="page-link" href="?page={pagination.next_num}"><i class="bi bi-chevron-right"></i></a></li>'
        content += '</ul></nav></div>'

    content += '</div>'
    return render_page('Сессии', 'sessions', content)


# ============== ADMINS ==============
@bp.route('/adminuser/')
@login_required
def admins_list():
    admins = AdminUser.query.order_by(AdminUser.created_at.desc()).all()

    content = '''
    <div class="page-header">
        <h1 class="page-title"><i class="bi bi-shield-lock me-2"></i>Администраторы</h1>
        <a href="/admin/adminuser/new" class="btn btn-primary"><i class="bi bi-plus-lg me-2"></i>Добавить</a>
    </div>
    <div class="card">
        <div class="card-header">
            <span class="card-title">Список администраторов</span>
        </div>
        <div class="table-responsive">
            <table class="table">
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Логин</th>
                        <th>Роль</th>
                        <th>Статус</th>
                        <th>Создан</th>
                        <th>Действия</th>
                    </tr>
                </thead>
                <tbody>
    '''

    role_labels = {'superadmin': 'Суперадмин', 'game_admin': 'Админ игры', 'quest_admin': 'Админ квеста'}
    for a in admins:
        status = '<span class="badge badge-success">Активен</span>' if a.is_active else '<span class="badge badge-danger">Отключён</span>'
        role = role_labels.get(a.role or 'superadmin', a.role or 'superadmin')
        date = a.created_at.strftime('%d.%m.%Y %H:%M') if a.created_at else '—'
        content += f'''
                    <tr>
                        <td><strong>#{a.id}</strong></td>
                        <td>{a.username}</td>
                        <td><span class="badge badge-info">{role}</span></td>
                        <td>{status}</td>
                        <td>{date}</td>
                        <td class="action-btns">
                            <a href="/admin/adminuser/edit/{a.id}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
                            <a href="/admin/adminuser/delete/{a.id}" class="btn btn-sm btn-danger" onclick="return confirm('Удалить администратора?')"><i class="bi bi-trash"></i></a>
                        </td>
                    </tr>
        '''

    content += '</tbody></table></div></div>'
    return render_page('Администраторы', 'admins', content)


@bp.route('/adminuser/new', methods=['GET', 'POST'])
@login_required
def admin_new():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role', 'quest_admin')
        if role not in ('superadmin', 'game_admin', 'quest_admin'):
            role = 'quest_admin'

        if AdminUser.query.filter_by(username=username).first():
            flash('Пользователь с таким логином уже существует!', 'danger')
        else:
            admin = AdminUser(username=username, is_active=True, role=role)
            admin.set_password(password)
            db.session.add(admin)
            db.session.commit()
            flash('Администратор создан!', 'success')
            return redirect(url_for('custom_admin.admins_list'))

    content = '''
    <div class="page-header">
        <h1 class="page-title"><i class="bi bi-plus-lg me-2"></i>Новый администратор</h1>
    </div>
    <div class="card">
        <div class="card-body">
            <form method="POST">
                <div class="row">
                    <div class="col-md-4 mb-3">
                        <label class="form-label">Логин</label>
                        <input type="text" name="username" class="form-control" required>
                    </div>
                    <div class="col-md-4 mb-3">
                        <label class="form-label">Пароль</label>
                        <input type="password" name="password" class="form-control" required>
                    </div>
                    <div class="col-md-4 mb-3">
                        <label class="form-label">Роль</label>
                        <select name="role" class="form-select">
                            <option value="superadmin">Суперадмин</option>
                            <option value="game_admin">Админ игры</option>
                            <option value="quest_admin" selected>Админ квеста</option>
                        </select>
                    </div>
                </div>
                <div class="d-flex gap-2">
                    <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-2"></i>Создать</button>
                    <a href="/admin/adminuser/" class="btn btn-outline-secondary">Отмена</a>
                </div>
            </form>
        </div>
    </div>
    '''
    return render_page('Новый администратор', 'admins', content)


@bp.route('/adminuser/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def admin_edit(id):
    admin = AdminUser.query.get_or_404(id)

    if request.method == 'POST':
        admin.username = request.form.get('username', admin.username)
        admin.is_active = request.form.get('is_active') == '1'
        role = request.form.get('role', admin.role or 'superadmin')
        if role in ('superadmin', 'game_admin', 'quest_admin'):
            admin.role = role
        new_password = request.form.get('password')
        if new_password:
            admin.set_password(new_password)
        db.session.commit()
        flash('Администратор обновлён!', 'success')
        return redirect(url_for('custom_admin.admins_list'))

    content = f'''
    <div class="page-header">
        <h1 class="page-title"><i class="bi bi-pencil me-2"></i>Редактирование: {admin.username}</h1>
    </div>
    <div class="card">
        <div class="card-body">
            <form method="POST">
                <div class="row">
                    <div class="col-md-6 mb-3">
                        <label class="form-label">Логин</label>
                        <input type="text" name="username" class="form-control" value="{admin.username}" required>
                    </div>
                    <div class="col-md-6 mb-3">
                        <label class="form-label">Новый пароль <small class="text-muted">(оставьте пустым, чтобы не менять)</small></label>
                        <input type="password" name="password" class="form-control">
                    </div>
                    <div class="col-md-4 mb-3">
                        <label class="form-label">Статус</label>
                        <select name="is_active" class="form-select">
                            <option value="1" {'selected' if admin.is_active else ''}>Активен</option>
                            <option value="0" {'' if admin.is_active else 'selected'}>Отключён</option>
                        </select>
                    </div>
                    <div class="col-md-4 mb-3">
                        <label class="form-label">Роль</label>
                        <select name="role" class="form-select">
                            <option value="superadmin" {'selected' if (admin.role or 'superadmin') == 'superadmin' else ''}>Суперадмин</option>
                            <option value="game_admin" {'selected' if admin.role == 'game_admin' else ''}>Админ игры</option>
                            <option value="quest_admin" {'selected' if admin.role == 'quest_admin' else ''}>Админ квеста</option>
                        </select>
                    </div>
                </div>
                <div class="d-flex gap-2">
                    <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-2"></i>Сохранить</button>
                    <a href="/admin/adminuser/" class="btn btn-outline-secondary">Отмена</a>
                </div>
            </form>
        </div>
    </div>
    '''
    return render_page('Редактирование администратора', 'admins', content)


@bp.route('/adminuser/delete/<int:id>')
@login_required
def admin_delete(id):
    admin = AdminUser.query.get_or_404(id)
    if admin.id == current_user.id:
        flash('Нельзя удалить самого себя!', 'danger')
    else:
        db.session.delete(admin)
        db.session.commit()
        flash('Администратор удалён!', 'success')
    return redirect(url_for('custom_admin.admins_list'))


# ============== GAME TEXTS ==============
SECTION_LABELS = {
    'game': 'Игра',
    'quest': 'Квест',
    'rules': 'Правила',
    'landing': 'Лендинг',
}


@bp.route('/texts/')
@login_required
def texts_list():
    section_filter = request.args.get('section', '')
    query = GameText.query

    if section_filter:
        query = query.filter_by(section=section_filter)

    texts = query.order_by(GameText.section, GameText.key).all()
    sections = db.session.query(GameText.section).distinct().order_by(GameText.section).all()
    section_list = [s[0] for s in sections]

    export_section = f'?section={section_filter}' if section_filter else ''
    content = f'''
    <div class="page-header">
        <h1 class="page-title"><i class="bi bi-fonts me-2"></i>Тексты игры и квеста</h1>
        <div class="d-flex gap-2">
            <a href="/admin/texts/export/csv{export_section}" class="btn btn-outline-success"><i class="bi bi-file-earmark-spreadsheet me-1"></i>CSV</a>
            <a href="/admin/texts/export/json{export_section}" class="btn btn-outline-primary"><i class="bi bi-filetype-json me-1"></i>JSON</a>
        </div>
    </div>
    <div class="card">
        <div class="card-header">
            <span class="card-title">Управление текстами</span>
            <div class="d-flex gap-2 align-items-center">
    '''

    # Section filter tabs
    active_class = 'btn-primary' if not section_filter else 'btn-outline-secondary'
    content += f'<a href="/admin/texts/" class="btn btn-sm {active_class}">Все</a>'
    for sec in section_list:
        label = SECTION_LABELS.get(sec, sec.title())
        active_class = 'btn-primary' if section_filter == sec else 'btn-outline-secondary'
        content += f'<a href="/admin/texts/?section={sec}" class="btn btn-sm {active_class}">{label}</a>'

    content += '''
            </div>
        </div>
        <div class="table-responsive">
            <table class="table">
                <thead>
                    <tr>
                        <th>Ключ</th>
                        <th>Раздел</th>
                        <th>Описание</th>
                        <th>Текст</th>
                        <th>Действия</th>
                    </tr>
                </thead>
                <tbody>
    '''

    for t in texts:
        sec_label = SECTION_LABELS.get(t.section, t.section)
        sec_class = {'game': 'info', 'quest': 'success', 'rules': 'warning', 'landing': 'primary'}.get(t.section, 'info')
        # Truncate long values for display
        display_val = t.value if len(t.value) <= 80 else t.value[:77] + '...'
        # Escape HTML
        import html
        display_val = html.escape(display_val)
        label_esc = html.escape(t.label)

        content += f'''
                    <tr>
                        <td><code style="background:#f1f5f9;padding:2px 8px;border-radius:4px;font-size:0.8rem;">{t.key}</code></td>
                        <td><span class="badge badge-{sec_class}">{sec_label}</span></td>
                        <td>{label_esc}</td>
                        <td style="max-width:300px;word-break:break-word;">{display_val}</td>
                        <td>
                            <a href="/admin/texts/edit/{t.id}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
                        </td>
                    </tr>
        '''

    content += '</tbody></table></div></div>'
    return render_page('Тексты', 'texts', content)


@bp.route('/texts/export/<fmt>')
@login_required
def texts_export(fmt):
    import csv
    import io
    import json as json_lib

    section_filter = request.args.get('section', '')
    query = GameText.query
    if section_filter:
        query = query.filter_by(section=section_filter)
    texts = query.order_by(GameText.section, GameText.key).all()

    suffix = f'_{section_filter}' if section_filter else ''

    if fmt == 'json':
        data = [
            {'key': t.key, 'section': t.section, 'label': t.label, 'value': t.value}
            for t in texts
        ]
        return Response(
            json_lib.dumps(data, ensure_ascii=False, indent=2),
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename=texts{suffix}.json'}
        )

    # CSV with UTF-8 BOM for Excel compatibility
    output = io.StringIO()
    output.write('\ufeff')  # BOM for Excel
    writer = csv.writer(output)
    writer.writerow(['key', 'section', 'label', 'value'])
    for t in texts:
        writer.writerow([t.key, t.section, t.label, t.value])

    return Response(
        output.getvalue(),
        mimetype='text/csv; charset=utf-8-sig',
        headers={'Content-Disposition': f'attachment; filename=texts{suffix}.csv'}
    )


@bp.route('/texts/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def text_edit(id):
    t = GameText.query.get_or_404(id)

    if request.method == 'POST':
        t.value = request.form.get('value', t.value)
        db.session.commit()
        flash(f'Текст "{t.label}" обновлён', 'success')
        return redirect(url_for('custom_admin.texts_list', section=t.section))

    import html
    value_esc = html.escape(t.value)
    label_esc = html.escape(t.label)
    sec_label = SECTION_LABELS.get(t.section, t.section)

    content = f'''
    <div class="page-header">
        <h1 class="page-title"><i class="bi bi-pencil me-2"></i>Редактирование текста</h1>
        <a href="/admin/texts/?section={t.section}" class="btn btn-outline-secondary"><i class="bi bi-arrow-left me-2"></i>Назад</a>
    </div>
    <div class="card">
        <div class="card-header">
            <span class="card-title">{label_esc}</span>
            <span class="badge badge-info">{sec_label}</span>
        </div>
        <div class="card-body">
            <form method="POST">
                <div class="mb-3">
                    <label class="form-label">Ключ</label>
                    <input type="text" class="form-control" value="{t.key}" disabled>
                    <small class="text-muted">Ключ используется в коде, менять нельзя</small>
                </div>
                <div class="mb-3">
                    <label class="form-label">Описание</label>
                    <input type="text" class="form-control" value="{label_esc}" disabled>
                </div>
                <div class="mb-4">
                    <label class="form-label">Текст</label>
                    <textarea class="form-control" name="value" rows="4" style="font-size:1rem;">{value_esc}</textarea>
                </div>
                <div class="d-flex gap-2">
                    <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-2"></i>Сохранить</button>
                    <a href="/admin/texts/?section={t.section}" class="btn btn-outline-secondary">Отмена</a>
                </div>
            </form>
        </div>
    </div>
    '''
    return render_page('Редактирование текста', 'texts', content)


# ============== ANALYTICS ==============
@bp.route('/analytics/')
@login_required
def analytics():
    from sqlalchemy import func, distinct
    from datetime import datetime, timedelta

    # ─── Summary Stats ───
    total_users = User.query.count()
    verified_users = User.query.filter_by(is_verified=True).count()

    # Users who played at least one game
    users_played = db.session.query(func.count(distinct(GameSession.user_id))).scalar() or 0

    # Total levels count
    total_levels = Level.query.filter_by(is_active=True).count()

    # Users who completed ALL levels (have progress for all active levels)
    users_completed_all = db.session.query(UserLevelProgress.user_id).filter(
        UserLevelProgress.completed_at.isnot(None)
    ).group_by(UserLevelProgress.user_id).having(
        func.count(UserLevelProgress.level_id) >= total_levels
    ).count() if total_levels > 0 else 0

    total_games = GameSession.query.count()
    won_games = GameSession.query.filter_by(is_won=True).count()
    win_rate = round((won_games / total_games * 100), 1) if total_games > 0 else 0

    # ─── Registrations by day (last 30 days) ───
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    registrations_by_day = db.session.query(
        func.date(User.created_at).label('date'),
        func.count(User.id).label('count')
    ).filter(User.created_at >= thirty_days_ago).group_by(
        func.date(User.created_at)
    ).order_by(func.date(User.created_at)).all()

    reg_labels = [r.date.strftime('%d.%m') if r.date else '' for r in registrations_by_day]
    reg_data = [r.count for r in registrations_by_day]

    # ─── Games by day (last 30 days) ───
    games_by_day = db.session.query(
        func.date(GameSession.created_at).label('date'),
        func.count(GameSession.id).label('count')
    ).filter(GameSession.created_at >= thirty_days_ago).group_by(
        func.date(GameSession.created_at)
    ).order_by(func.date(GameSession.created_at)).all()

    games_labels = [g.date.strftime('%d.%m') if g.date else '' for g in games_by_day]
    games_data = [g.count for g in games_by_day]

    # ─── Level completion funnel ───
    levels = Level.query.filter_by(is_active=True).order_by(Level.order).all()
    level_stats = []
    for lvl in levels:
        completed = UserLevelProgress.query.filter(
            UserLevelProgress.level_id == lvl.id,
            UserLevelProgress.completed_at.isnot(None)
        ).count()
        attempts = db.session.query(func.sum(UserLevelProgress.attempts_count)).filter(
            UserLevelProgress.level_id == lvl.id
        ).scalar() or 0
        level_stats.append({
            'name': lvl.name,
            'order': lvl.order,
            'completed': completed,
            'attempts': attempts
        })

    funnel_labels = [l['name'] for l in level_stats]
    funnel_data = [l['completed'] for l in level_stats]

    # ─── Top 10 players ───
    top_players = User.query.filter(
        User.is_verified == True,
        User.total_score > 0
    ).order_by(User.total_score.desc()).limit(10).all()

    # ─── Recent activity (last 7 days) ───
    seven_days_ago = datetime.utcnow() - timedelta(days=7)
    recent_registrations = User.query.filter(User.created_at >= seven_days_ago).count()
    recent_games = GameSession.query.filter(GameSession.created_at >= seven_days_ago).count()
    recent_completions = UserLevelProgress.query.filter(
        UserLevelProgress.completed_at >= seven_days_ago
    ).count()

    # ─── Region Stats ───
    moscow_total = User.query.filter(User.city == 'moscow').count()
    moscow_verified = User.query.filter(User.city == 'moscow', User.is_verified == True).count()
    moscow_played = db.session.query(func.count(distinct(GameSession.user_id))).join(
        User, GameSession.user_id == User.id
    ).filter(User.city == 'moscow').scalar() or 0
    moscow_completed_all = db.session.query(UserLevelProgress.user_id).join(
        User, UserLevelProgress.user_id == User.id
    ).filter(
        User.city == 'moscow',
        UserLevelProgress.completed_at.isnot(None)
    ).group_by(UserLevelProgress.user_id).having(
        func.count(UserLevelProgress.level_id) >= total_levels
    ).count() if total_levels > 0 else 0

    region_total = User.query.filter(User.city == 'region').count()
    region_verified = User.query.filter(User.city == 'region', User.is_verified == True).count()
    region_played = db.session.query(func.count(distinct(GameSession.user_id))).join(
        User, GameSession.user_id == User.id
    ).filter(User.city == 'region').scalar() or 0
    region_completed_all = db.session.query(UserLevelProgress.user_id).join(
        User, UserLevelProgress.user_id == User.id
    ).filter(
        User.city == 'region',
        UserLevelProgress.completed_at.isnot(None)
    ).group_by(UserLevelProgress.user_id).having(
        func.count(UserLevelProgress.level_id) >= total_levels
    ).count() if total_levels > 0 else 0

    # Registrations by day per region (last 30 days)
    reg_moscow_by_day = db.session.query(
        func.date(User.created_at).label('date'),
        func.count(User.id).label('count')
    ).filter(User.created_at >= thirty_days_ago, User.city == 'moscow').group_by(
        func.date(User.created_at)
    ).order_by(func.date(User.created_at)).all()

    reg_region_by_day = db.session.query(
        func.date(User.created_at).label('date'),
        func.count(User.id).label('count')
    ).filter(User.created_at >= thirty_days_ago, User.city == 'region').group_by(
        func.date(User.created_at)
    ).order_by(func.date(User.created_at)).all()

    # Merge dates for region registrations chart
    all_reg_dates = sorted(set(
        [r.date for r in reg_moscow_by_day if r.date] +
        [r.date for r in reg_region_by_day if r.date]
    ))
    moscow_day_map = {r.date: r.count for r in reg_moscow_by_day if r.date}
    region_day_map = {r.date: r.count for r in reg_region_by_day if r.date}
    region_reg_labels = [d.strftime('%d.%m') for d in all_reg_dates]
    region_reg_moscow_data = [moscow_day_map.get(d, 0) for d in all_reg_dates]
    region_reg_region_data = [region_day_map.get(d, 0) for d in all_reg_dates]

    # Top 10 per region
    top_moscow = User.query.filter(
        User.is_verified == True, User.total_score > 0, User.city == 'moscow'
    ).order_by(User.total_score.desc()).limit(10).all()
    top_region = User.query.filter(
        User.is_verified == True, User.total_score > 0, User.city == 'region'
    ).order_by(User.total_score.desc()).limit(10).all()

    # Full region breakdown (all city_names for "region" users)
    city_name_stats = db.session.query(
        User.city_name,
        func.count(User.id).label('count')
    ).filter(
        User.city == 'region',
        User.city_name.isnot(None),
        User.city_name != ''
    ).group_by(User.city_name).order_by(func.count(User.id).desc()).all()

    # Moscow breakdown (Москва vs МО)
    moscow_name_stats = db.session.query(
        User.city_name,
        func.count(User.id).label('count')
    ).filter(
        User.city == 'moscow',
        User.city_name.isnot(None),
        User.city_name != ''
    ).group_by(User.city_name).order_by(func.count(User.id).desc()).all()

    # Get or create share token
    from app.models import AnalyticsShare
    share = AnalyticsShare.query.first()
    if not share:
        import secrets
        share = AnalyticsShare(token=secrets.token_urlsafe(16), password_hash='')
        db.session.add(share)
        db.session.commit()

    content = f'''
    <div class="page-header">
        <h1 class="page-title"><i class="bi bi-bar-chart-line me-2"></i>Аналитика</h1>
        <div class="d-flex gap-2 flex-wrap">
            <a href="/admin/analytics/export/csv" class="btn btn-outline-secondary"><i class="bi bi-filetype-csv me-2"></i>CSV</a>
            <a href="/admin/analytics/export/xlsx" class="btn btn-outline-secondary"><i class="bi bi-file-earmark-excel me-2"></i>Excel</a>
            <a href="/admin/analytics/export/json" class="btn btn-outline-secondary"><i class="bi bi-filetype-json me-2"></i>JSON</a>
            <button class="btn btn-primary" data-bs-toggle="modal" data-bs-target="#shareModal"><i class="bi bi-share me-2"></i>Поделиться</button>
        </div>
    </div>

    <!-- Share Modal -->
    <div class="modal fade" id="shareModal" tabindex="-1">
        <div class="modal-dialog">
            <div class="modal-content">
                <div class="modal-header">
                    <h5 class="modal-title"><i class="bi bi-share me-2"></i>Поделиться аналитикой</h5>
                    <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
                </div>
                <form method="POST" action="/admin/analytics/share">
                    <div class="modal-body">
                        <p class="text-muted">Создайте защищённую ссылку на аналитику. Получатель сможет просмотреть графики и скачать PDF.</p>
                        <div class="mb-3">
                            <label class="form-label">Пароль для доступа</label>
                            <input type="password" name="share_password" class="form-control" placeholder="Введите пароль" required>
                        </div>
                        <div class="alert alert-info">
                            <i class="bi bi-info-circle me-2"></i>
                            <strong>Публичная ссылка:</strong><br>
                            <code id="shareUrl">{request.host_url}admin/public/{share.token}</code>
                            <button type="button" class="btn btn-sm btn-outline-secondary ms-2" onclick="navigator.clipboard.writeText(document.getElementById('shareUrl').innerText)">
                                <i class="bi bi-clipboard"></i>
                            </button>
                        </div>
                    </div>
                    <div class="modal-footer">
                        <button type="button" class="btn btn-outline-secondary" data-bs-dismiss="modal">Отмена</button>
                        <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-2"></i>Сохранить пароль</button>
                    </div>
                </form>
            </div>
        </div>
    </div>

    <!-- Tab Navigation -->
    <ul class="nav nav-tabs mb-4" role="tablist">
        <li class="nav-item">
            <a class="nav-link active" id="game-tab" data-bs-toggle="tab" href="#game-analytics" role="tab">
                <i class="bi bi-controller me-2"></i>Игра
            </a>
        </li>
        <li class="nav-item">
            <a class="nav-link" id="quest-tab" data-bs-toggle="tab" href="#quest-analytics" role="tab">
                <i class="bi bi-map me-2"></i>Квест
            </a>
        </li>
        <li class="nav-item">
            <a class="nav-link" id="regions-tab" data-bs-toggle="tab" href="#regions-analytics" role="tab">
                <i class="bi bi-geo-alt me-2"></i>Регионы
            </a>
        </li>
    </ul>

    <div class="tab-content">
    <div class="tab-pane fade show active" id="game-analytics" role="tabpanel">

    <!-- Summary Cards -->
    <div class="row mb-4">
        <div class="col-md-3 col-sm-6 mb-3">
            <div class="card h-100" style="border-left: 4px solid var(--primary);">
                <div class="card-body">
                    <div class="d-flex justify-content-between align-items-start">
                        <div>
                            <div class="text-muted small text-uppercase mb-1">Всего пользователей</div>
                            <div style="font-size: 2rem; font-weight: 700; color: var(--dark);">{total_users:,}</div>
                            <div class="text-muted small">Подтверждённых: {verified_users:,}</div>
                        </div>
                        <div style="width: 48px; height: 48px; background: rgba(99, 102, 241, 0.1); border-radius: 12px; display: flex; align-items: center; justify-content: center;">
                            <i class="bi bi-people" style="font-size: 1.5rem; color: var(--primary);"></i>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        <div class="col-md-3 col-sm-6 mb-3">
            <div class="card h-100" style="border-left: 4px solid var(--success);">
                <div class="card-body">
                    <div class="d-flex justify-content-between align-items-start">
                        <div>
                            <div class="text-muted small text-uppercase mb-1">Играли</div>
                            <div style="font-size: 2rem; font-weight: 700; color: var(--dark);">{users_played:,}</div>
                            <div class="text-muted small">{round(users_played/total_users*100, 1) if total_users > 0 else 0}% от всех</div>
                        </div>
                        <div style="width: 48px; height: 48px; background: rgba(16, 185, 129, 0.1); border-radius: 12px; display: flex; align-items: center; justify-content: center;">
                            <i class="bi bi-joystick" style="font-size: 1.5rem; color: var(--success);"></i>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        <div class="col-md-3 col-sm-6 mb-3">
            <div class="card h-100" style="border-left: 4px solid var(--warning);">
                <div class="card-body">
                    <div class="d-flex justify-content-between align-items-start">
                        <div>
                            <div class="text-muted small text-uppercase mb-1">Прошли всё</div>
                            <div style="font-size: 2rem; font-weight: 700; color: var(--dark);">{users_completed_all:,}</div>
                            <div class="text-muted small">Все {total_levels} уровней</div>
                        </div>
                        <div style="width: 48px; height: 48px; background: rgba(245, 158, 11, 0.1); border-radius: 12px; display: flex; align-items: center; justify-content: center;">
                            <i class="bi bi-trophy" style="font-size: 1.5rem; color: var(--warning);"></i>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        <div class="col-md-3 col-sm-6 mb-3">
            <div class="card h-100" style="border-left: 4px solid var(--info);">
                <div class="card-body">
                    <div class="d-flex justify-content-between align-items-start">
                        <div>
                            <div class="text-muted small text-uppercase mb-1">Всего игр</div>
                            <div style="font-size: 2rem; font-weight: 700; color: var(--dark);">{total_games:,}</div>
                            <div class="text-muted small">Win rate: {win_rate}%</div>
                        </div>
                        <div style="width: 48px; height: 48px; background: rgba(6, 182, 212, 0.1); border-radius: 12px; display: flex; align-items: center; justify-content: center;">
                            <i class="bi bi-controller" style="font-size: 1.5rem; color: var(--info);"></i>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <!-- Recent Activity Cards -->
    <div class="row mb-4">
        <div class="col-12">
            <div class="card">
                <div class="card-header">
                    <span class="card-title"><i class="bi bi-clock-history me-2"></i>За последние 7 дней</span>
                </div>
                <div class="card-body">
                    <div class="row text-center">
                        <div class="col-md-4">
                            <div style="font-size: 2.5rem; font-weight: 700; color: var(--primary);">{recent_registrations}</div>
                            <div class="text-muted">Новых регистраций</div>
                        </div>
                        <div class="col-md-4">
                            <div style="font-size: 2.5rem; font-weight: 700; color: var(--success);">{recent_games}</div>
                            <div class="text-muted">Сыграно игр</div>
                        </div>
                        <div class="col-md-4">
                            <div style="font-size: 2.5rem; font-weight: 700; color: var(--warning);">{recent_completions}</div>
                            <div class="text-muted">Уровней пройдено</div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <!-- Charts Row -->
    <div class="row mb-4">
        <div class="col-md-6 mb-3">
            <div class="card h-100">
                <div class="card-header">
                    <span class="card-title"><i class="bi bi-person-plus me-2"></i>Регистрации (30 дней)</span>
                </div>
                <div class="card-body">
                    <canvas id="registrationsChart" height="200"></canvas>
                </div>
            </div>
        </div>
        <div class="col-md-6 mb-3">
            <div class="card h-100">
                <div class="card-header">
                    <span class="card-title"><i class="bi bi-controller me-2"></i>Игры (30 дней)</span>
                </div>
                <div class="card-body">
                    <canvas id="gamesChart" height="200"></canvas>
                </div>
            </div>
        </div>
    </div>

    <!-- Level Funnel -->
    <div class="row mb-4">
        <div class="col-12">
            <div class="card">
                <div class="card-header">
                    <span class="card-title"><i class="bi bi-funnel me-2"></i>Воронка прохождения уровней</span>
                </div>
                <div class="card-body">
                    <canvas id="funnelChart" height="100"></canvas>
                </div>
            </div>
        </div>
    </div>

    <!-- Level Stats Table -->
    <div class="row mb-4">
        <div class="col-md-6 mb-3">
            <div class="card h-100">
                <div class="card-header">
                    <span class="card-title"><i class="bi bi-layers me-2"></i>Статистика по уровням</span>
                </div>
                <div class="table-responsive">
                    <table class="table">
                        <thead>
                            <tr>
                                <th>Уровень</th>
                                <th>Прошли</th>
                                <th>Попыток</th>
                                <th>Конверсия</th>
                            </tr>
                        </thead>
                        <tbody>
    '''

    for ls in level_stats:
        conv = round(ls['completed'] / ls['attempts'] * 100, 1) if ls['attempts'] > 0 else 0
        conv_badge = 'badge-success' if conv >= 50 else 'badge-warning' if conv >= 25 else 'badge-danger'
        content += f'''
                            <tr>
                                <td><strong>{ls['name']}</strong></td>
                                <td>{ls['completed']:,}</td>
                                <td>{ls['attempts']:,}</td>
                                <td><span class="badge {conv_badge}">{conv}%</span></td>
                            </tr>
        '''

    content += '''
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
        <div class="col-md-6 mb-3">
            <div class="card h-100">
                <div class="card-header">
                    <span class="card-title"><i class="bi bi-trophy me-2"></i>Топ-10 игроков</span>
                </div>
                <div class="table-responsive">
                    <table class="table">
                        <thead>
                            <tr>
                                <th>#</th>
                                <th>Игрок</th>
                                <th>Очки</th>
                            </tr>
                        </thead>
                        <tbody>
    '''

    for idx, player in enumerate(top_players, 1):
        medal = '🥇' if idx == 1 else '🥈' if idx == 2 else '🥉' if idx == 3 else str(idx)
        content += f'''
                            <tr>
                                <td>{medal}</td>
                                <td><strong>{player.username}</strong></td>
                                <td>{player.total_score:,}</td>
                            </tr>
        '''

    content += '''
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script>
        // Registrations Chart
        new Chart(document.getElementById('registrationsChart'), {
            type: 'line',
            data: {
                labels: ''' + str(reg_labels) + ''',
                datasets: [{
                    label: 'Регистрации',
                    data: ''' + str(reg_data) + ''',
                    borderColor: '#6366f1',
                    backgroundColor: 'rgba(99, 102, 241, 0.1)',
                    fill: true,
                    tension: 0.4
                }]
            },
            options: {
                responsive: true,
                plugins: { legend: { display: false } },
                scales: {
                    y: { beginAtZero: true, ticks: { precision: 0 } }
                }
            }
        });

        // Games Chart
        new Chart(document.getElementById('gamesChart'), {
            type: 'line',
            data: {
                labels: ''' + str(games_labels) + ''',
                datasets: [{
                    label: 'Игры',
                    data: ''' + str(games_data) + ''',
                    borderColor: '#10b981',
                    backgroundColor: 'rgba(16, 185, 129, 0.1)',
                    fill: true,
                    tension: 0.4
                }]
            },
            options: {
                responsive: true,
                plugins: { legend: { display: false } },
                scales: {
                    y: { beginAtZero: true, ticks: { precision: 0 } }
                }
            }
        });

        // Funnel Chart
        new Chart(document.getElementById('funnelChart'), {
            type: 'bar',
            data: {
                labels: ''' + str(funnel_labels) + ''',
                datasets: [{
                    label: 'Прошли уровень',
                    data: ''' + str(funnel_data) + ''',
                    backgroundColor: [
                        'rgba(99, 102, 241, 0.8)',
                        'rgba(16, 185, 129, 0.8)',
                        'rgba(245, 158, 11, 0.8)',
                        'rgba(239, 68, 68, 0.8)',
                        'rgba(6, 182, 212, 0.8)',
                        'rgba(168, 85, 247, 0.8)'
                    ],
                    borderRadius: 8
                }]
            },
            options: {
                responsive: true,
                plugins: { legend: { display: false } },
                scales: {
                    y: { beginAtZero: true, ticks: { precision: 0 } }
                }
            }
        });
    </script>
    '''

    # Close game tab pane
    content += '</div>'  # close #game-analytics

    # ─── Quest Analytics Tab ───
    from app.models import QuestPage, QuestProgress, PromoCodePool, PromoCode

    quest_participants = db.session.query(func.count(distinct(QuestProgress.user_id))).scalar() or 0
    quest_total_answers = QuestProgress.query.count()
    quest_correct = QuestProgress.query.filter_by(is_correct=True).count()
    quest_skipped = QuestProgress.query.filter_by(is_skipped=True).count()
    quest_total_pages = QuestPage.query.filter_by(is_active=True).count()
    avg_quest_score = db.session.query(func.avg(User.quest_score)).filter(User.quest_score > 0).scalar() or 0

    # Completion rate (users who answered all pages)
    quest_completed_users = 0
    if quest_total_pages > 0:
        quest_completed_users = db.session.query(QuestProgress.user_id).group_by(
            QuestProgress.user_id
        ).having(func.count(QuestProgress.id) >= quest_total_pages).count()

    completion_rate = round(quest_completed_users / quest_participants * 100, 1) if quest_participants > 0 else 0

    # Promo stats
    promo_pools = PromoCodePool.query.order_by(PromoCodePool.min_score.desc()).all()
    total_promos_issued = sum(p.used_codes for p in promo_pools)

    # Registration source breakdown
    source_game = User.query.filter_by(registration_source='game').count()
    source_quest = User.query.filter_by(registration_source='quest').count()
    source_transferred = User.query.filter_by(registration_source='transferred').count()

    # Quest participation by day (last 30 days)
    quest_by_day = db.session.query(
        func.date(QuestProgress.scanned_at).label('date'),
        func.count(QuestProgress.id).label('count')
    ).filter(QuestProgress.scanned_at >= thirty_days_ago).group_by(
        func.date(QuestProgress.scanned_at)
    ).order_by(func.date(QuestProgress.scanned_at)).all()

    quest_day_labels = [q.date.strftime('%d.%m') if q.date else '' for q in quest_by_day]
    quest_day_data = [q.count for q in quest_by_day]

    content += f'''
    <div class="tab-pane fade" id="quest-analytics" role="tabpanel">

    <!-- Quest Summary Cards -->
    <div class="row mb-4">
        <div class="col-md-3 col-sm-6 mb-3">
            <div class="card h-100" style="border-left: 4px solid var(--primary);">
                <div class="card-body">
                    <div class="text-muted small text-uppercase mb-1">Участников квеста</div>
                    <div style="font-size: 2rem; font-weight: 700; color: var(--dark);">{quest_participants}</div>
                    <div class="text-muted small">Прошли полностью: {quest_completed_users}</div>
                </div>
            </div>
        </div>
        <div class="col-md-3 col-sm-6 mb-3">
            <div class="card h-100" style="border-left: 4px solid var(--success);">
                <div class="card-body">
                    <div class="text-muted small text-uppercase mb-1">Средний балл</div>
                    <div style="font-size: 2rem; font-weight: 700; color: var(--dark);">{avg_quest_score:.0f}</div>
                    <div class="text-muted small">Процент прохождения: {completion_rate}%</div>
                </div>
            </div>
        </div>
        <div class="col-md-3 col-sm-6 mb-3">
            <div class="card h-100" style="border-left: 4px solid var(--warning);">
                <div class="card-body">
                    <div class="text-muted small text-uppercase mb-1">Всего ответов</div>
                    <div style="font-size: 2rem; font-weight: 700; color: var(--dark);">{quest_total_answers}</div>
                    <div class="text-muted small">Верных: {quest_correct} | Пропущено: {quest_skipped}</div>
                </div>
            </div>
        </div>
        <div class="col-md-3 col-sm-6 mb-3">
            <div class="card h-100" style="border-left: 4px solid var(--info);">
                <div class="card-body">
                    <div class="text-muted small text-uppercase mb-1">Промокодов выдано</div>
                    <div style="font-size: 2rem; font-weight: 700; color: var(--dark);">{total_promos_issued}</div>
                    <div class="text-muted small">Пулов: {len(promo_pools)}</div>
                </div>
            </div>
        </div>
    </div>

    <!-- Registration Source Breakdown -->
    <div class="row mb-4">
        <div class="col-md-6 mb-3">
            <div class="card h-100">
                <div class="card-header"><span class="card-title"><i class="bi bi-diagram-3 me-2"></i>Источники регистрации</span></div>
                <div class="card-body">
                    <canvas id="sourceChart" height="200"></canvas>
                </div>
            </div>
        </div>
        <div class="col-md-6 mb-3">
            <div class="card h-100">
                <div class="card-header"><span class="card-title"><i class="bi bi-ticket-perforated me-2"></i>Промокоды по уровням</span></div>
                <div class="card-body">
                    <table class="table">
                        <thead><tr><th>Уровень</th><th>Мин. очков</th><th>Выдано</th><th>Осталось</th><th>Статус</th></tr></thead>
                        <tbody>
    '''

    for pool in promo_pools:
        remaining = pool.remaining_codes
        status_class = 'badge-danger' if pool.is_low else 'badge-success'
        status_text = 'Мало!' if pool.is_low else 'ОК'
        content += f'''
                            <tr>
                                <td><strong>{pool.name}</strong></td>
                                <td>{pool.min_score}</td>
                                <td>{pool.used_codes}</td>
                                <td>{remaining}</td>
                                <td><span class="badge {status_class}">{status_text}</span></td>
                            </tr>
        '''

    content += f'''
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>

    <!-- Quest Activity Chart -->
    <div class="card mb-4">
        <div class="card-header"><span class="card-title"><i class="bi bi-graph-up me-2"></i>Активность квеста (30 дней)</span></div>
        <div class="card-body">
            <canvas id="questActivityChart" height="100"></canvas>
        </div>
    </div>

    <script>
        // Source breakdown pie chart
        new Chart(document.getElementById('sourceChart'), {{
            type: 'doughnut',
            data: {{
                labels: ['Игра', 'Квест', 'Перетёкшие'],
                datasets: [{{
                    data: [{source_game}, {source_quest}, {source_transferred}],
                    backgroundColor: ['rgba(99, 102, 241, 0.8)', 'rgba(16, 185, 129, 0.8)', 'rgba(245, 158, 11, 0.8)'],
                    borderWidth: 0
                }}]
            }},
            options: {{
                responsive: true,
                plugins: {{
                    legend: {{ position: 'bottom' }}
                }}
            }}
        }});

        // Quest activity line chart
        new Chart(document.getElementById('questActivityChart'), {{
            type: 'line',
            data: {{
                labels: {quest_day_labels},
                datasets: [{{
                    label: 'Ответы в квесте',
                    data: {quest_day_data},
                    borderColor: 'rgba(16, 185, 129, 1)',
                    backgroundColor: 'rgba(16, 185, 129, 0.1)',
                    fill: true,
                    tension: 0.3,
                    pointRadius: 3
                }}]
            }},
            options: {{
                responsive: true,
                plugins: {{ legend: {{ display: false }} }},
                scales: {{
                    y: {{ beginAtZero: true, ticks: {{ precision: 0 }} }}
                }}
            }}
        }});
    </script>

    </div>
    '''

    # ─── Regions Analytics Tab ───
    from markupsafe import escape

    city_name_rows = ''
    for idx, cs in enumerate(city_name_stats, 1):
        pct = round(cs.count / region_total * 100, 1) if region_total > 0 else 0
        city_name_rows += f'''
                            <tr>
                                <td class="text-muted">{idx}</td>
                                <td><strong>{escape(cs.city_name)}</strong></td>
                                <td>{cs.count:,}</td>
                                <td style="width:35%;"><div class="progress" style="height:6px;"><div class="progress-bar" style="width:{pct}%;background:var(--success);"></div></div></td>
                                <td>{pct}%</td>
                            </tr>'''

    moscow_name_rows = ''
    for cs in moscow_name_stats:
        pct = round(cs.count / moscow_total * 100, 1) if moscow_total > 0 else 0
        moscow_name_rows += f'''
                            <tr>
                                <td><strong>{escape(cs.city_name)}</strong></td>
                                <td>{cs.count:,}</td>
                                <td style="width:40%;"><div class="progress" style="height:6px;"><div class="progress-bar" style="width:{pct}%;background:var(--primary);"></div></div></td>
                                <td>{pct}%</td>
                            </tr>'''

    top_moscow_rows = ''
    for idx, player in enumerate(top_moscow, 1):
        medal = '\U0001f947' if idx == 1 else '\U0001f948' if idx == 2 else '\U0001f949' if idx == 3 else str(idx)
        top_moscow_rows += f'''
                            <tr>
                                <td>{medal}</td>
                                <td><strong>{player.username}</strong></td>
                                <td>{player.total_score:,}</td>
                            </tr>'''

    top_region_rows = ''
    for idx, player in enumerate(top_region, 1):
        medal = '\U0001f947' if idx == 1 else '\U0001f948' if idx == 2 else '\U0001f949' if idx == 3 else str(idx)
        top_region_rows += f'''
                            <tr>
                                <td>{medal}</td>
                                <td><strong>{player.username}</strong></td>
                                <td>{player.total_score:,}</td>
                            </tr>'''

    content += f'''
    <div class="tab-pane fade" id="regions-analytics" role="tabpanel">

    <!-- Region Summary Cards -->
    <div class="row mb-4">
        <div class="col-md-6 mb-3">
            <div class="card h-100" style="border-left: 4px solid var(--primary);">
                <div class="card-body">
                    <div class="d-flex justify-content-between align-items-start">
                        <div>
                            <div class="text-muted small text-uppercase mb-1">Москва и МО</div>
                            <div style="font-size: 2rem; font-weight: 700; color: var(--dark);">{moscow_total:,}</div>
                            <div class="text-muted small">
                                Подтверждённых: {moscow_verified:,} |
                                Играли: {moscow_played:,} |
                                Прошли всё: {moscow_completed_all:,}
                            </div>
                        </div>
                        <div style="width: 48px; height: 48px; background: rgba(99, 102, 241, 0.1); border-radius: 12px; display: flex; align-items: center; justify-content: center;">
                            <i class="bi bi-building" style="font-size: 1.5rem; color: var(--primary);"></i>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        <div class="col-md-6 mb-3">
            <div class="card h-100" style="border-left: 4px solid var(--success);">
                <div class="card-body">
                    <div class="d-flex justify-content-between align-items-start">
                        <div>
                            <div class="text-muted small text-uppercase mb-1">Регионы</div>
                            <div style="font-size: 2rem; font-weight: 700; color: var(--dark);">{region_total:,}</div>
                            <div class="text-muted small">
                                Подтверждённых: {region_verified:,} |
                                Играли: {region_played:,} |
                                Прошли всё: {region_completed_all:,}
                            </div>
                        </div>
                        <div style="width: 48px; height: 48px; background: rgba(16, 185, 129, 0.1); border-radius: 12px; display: flex; align-items: center; justify-content: center;">
                            <i class="bi bi-geo-alt" style="font-size: 1.5rem; color: var(--success);"></i>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <!-- Region Distribution + Registrations Chart -->
    <div class="row mb-4">
        <div class="col-md-4 mb-3">
            <div class="card h-100">
                <div class="card-header">
                    <span class="card-title"><i class="bi bi-pie-chart me-2"></i>Распределение</span>
                </div>
                <div class="card-body">
                    <canvas id="regionPieChart" height="250"></canvas>
                </div>
            </div>
        </div>
        <div class="col-md-8 mb-3">
            <div class="card h-100">
                <div class="card-header">
                    <span class="card-title"><i class="bi bi-graph-up me-2"></i>Регистрации по дням (30 дней)</span>
                </div>
                <div class="card-body">
                    <canvas id="regionRegChart" height="200"></canvas>
                </div>
            </div>
        </div>
    </div>

    <!-- Detailed region breakdown -->
    <div class="row mb-4">
        <div class="col-md-4 mb-3">
            <div class="card h-100">
                <div class="card-header">
                    <span class="card-title"><i class="bi bi-building me-2"></i>Москва и МО ({moscow_total:,})</span>
                </div>
                <div class="table-responsive" style="max-height:500px;overflow-y:auto;">
                    <table class="table">
                        <thead><tr><th>Регион</th><th>Кол-во</th><th></th><th>%</th></tr></thead>
                        <tbody>
                            {moscow_name_rows if moscow_name_rows else '<tr><td colspan="4" class="text-center text-muted">Нет данных</td></tr>'}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
        <div class="col-md-8 mb-3">
            <div class="card h-100">
                <div class="card-header d-flex justify-content-between align-items-center">
                    <span class="card-title"><i class="bi bi-geo-alt me-2"></i>Регионы ({region_total:,})</span>
                    <span class="text-muted small">{len(city_name_stats)} регионов</span>
                </div>
                <div class="table-responsive" style="max-height:500px;overflow-y:auto;">
                    <table class="table">
                        <thead><tr><th>#</th><th>Регион</th><th>Кол-во</th><th></th><th>%</th></tr></thead>
                        <tbody>
                            {city_name_rows if city_name_rows else '<tr><td colspan="5" class="text-center text-muted">Нет данных</td></tr>'}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>

    <!-- Top players per region -->
    <div class="row mb-4">
        <div class="col-md-6 mb-3">
            <div class="card h-100">
                <div class="card-header">
                    <span class="card-title"><i class="bi bi-trophy me-2"></i>Топ-10 Москва и МО</span>
                </div>
                <div class="table-responsive">
                    <table class="table">
                        <thead><tr><th>#</th><th>Игрок</th><th>Очки</th></tr></thead>
                        <tbody>
                            {top_moscow_rows if top_moscow_rows else '<tr><td colspan="3" class="text-center text-muted">Нет данных</td></tr>'}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
        <div class="col-md-6 mb-3">
            <div class="card h-100">
                <div class="card-header">
                    <span class="card-title"><i class="bi bi-trophy me-2"></i>Топ-10 Регионы</span>
                </div>
                <div class="table-responsive">
                    <table class="table">
                        <thead><tr><th>#</th><th>Игрок</th><th>Очки</th></tr></thead>
                        <tbody>
                            {top_region_rows if top_region_rows else '<tr><td colspan="3" class="text-center text-muted">Нет данных</td></tr>'}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>

    <script>
        // Region pie chart
        new Chart(document.getElementById('regionPieChart'), {{
            type: 'doughnut',
            data: {{
                labels: ['Москва и МО', 'Регионы'],
                datasets: [{{
                    data: [{moscow_total}, {region_total}],
                    backgroundColor: ['rgba(99, 102, 241, 0.8)', 'rgba(16, 185, 129, 0.8)'],
                    borderWidth: 0
                }}]
            }},
            options: {{
                responsive: true,
                plugins: {{
                    legend: {{ position: 'bottom' }}
                }}
            }}
        }});

        // Region registrations stacked chart
        new Chart(document.getElementById('regionRegChart'), {{
            type: 'bar',
            data: {{
                labels: {region_reg_labels},
                datasets: [
                    {{
                        label: 'Москва и МО',
                        data: {region_reg_moscow_data},
                        backgroundColor: 'rgba(99, 102, 241, 0.7)',
                        borderRadius: 4
                    }},
                    {{
                        label: 'Регионы',
                        data: {region_reg_region_data},
                        backgroundColor: 'rgba(16, 185, 129, 0.7)',
                        borderRadius: 4
                    }}
                ]
            }},
            options: {{
                responsive: true,
                plugins: {{ legend: {{ position: 'top' }} }},
                scales: {{
                    x: {{ stacked: true, grid: {{ display: false }} }},
                    y: {{ stacked: true, beginAtZero: true, ticks: {{ precision: 0 }} }}
                }}
            }}
        }});
    </script>

    </div>
    '''

    # Close tab-content
    content += '</div>'

    return render_page('Аналитика', 'analytics', content)


def get_analytics_data():
    """Helper function to gather all analytics data"""
    from sqlalchemy import func, distinct
    from datetime import datetime, timedelta

    total_users = User.query.count()
    verified_users = User.query.filter_by(is_verified=True).count()
    users_played = db.session.query(func.count(distinct(GameSession.user_id))).scalar() or 0
    total_levels = Level.query.filter_by(is_active=True).count()

    users_completed_all = db.session.query(UserLevelProgress.user_id).filter(
        UserLevelProgress.completed_at.isnot(None)
    ).group_by(UserLevelProgress.user_id).having(
        func.count(UserLevelProgress.level_id) >= total_levels
    ).count() if total_levels > 0 else 0

    total_games = GameSession.query.count()
    won_games = GameSession.query.filter_by(is_won=True).count()

    # Level stats
    levels = Level.query.filter_by(is_active=True).order_by(Level.order).all()
    level_stats = []
    for lvl in levels:
        completed = UserLevelProgress.query.filter(
            UserLevelProgress.level_id == lvl.id,
            UserLevelProgress.completed_at.isnot(None)
        ).count()
        attempts = db.session.query(func.sum(UserLevelProgress.attempts_count)).filter(
            UserLevelProgress.level_id == lvl.id
        ).scalar() or 0
        level_stats.append({
            'level_name': lvl.name,
            'order': lvl.order,
            'completed': completed,
            'attempts': int(attempts),
            'conversion': round(completed / attempts * 100, 2) if attempts > 0 else 0
        })

    # Charts data (last 30 days)
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)

    registrations_by_day = db.session.query(
        func.date(User.created_at).label('date'),
        func.count(User.id).label('count')
    ).filter(User.created_at >= thirty_days_ago).group_by(
        func.date(User.created_at)
    ).order_by(func.date(User.created_at)).all()

    games_by_day = db.session.query(
        func.date(GameSession.created_at).label('date'),
        func.count(GameSession.id).label('count')
    ).filter(GameSession.created_at >= thirty_days_ago).group_by(
        func.date(GameSession.created_at)
    ).order_by(func.date(GameSession.created_at)).all()

    # Top players
    top_players = User.query.filter(
        User.is_verified == True,
        User.total_score > 0
    ).order_by(User.total_score.desc()).limit(100).all()

    # Recent activity (7 days)
    seven_days_ago = datetime.utcnow() - timedelta(days=7)
    recent_registrations = User.query.filter(User.created_at >= seven_days_ago).count()
    recent_games = GameSession.query.filter(GameSession.created_at >= seven_days_ago).count()
    recent_completions = UserLevelProgress.query.filter(
        UserLevelProgress.completed_at >= seven_days_ago
    ).count()

    # Region stats
    moscow_total = User.query.filter(User.city == 'moscow').count()
    moscow_verified = User.query.filter(User.city == 'moscow', User.is_verified == True).count()
    moscow_played = db.session.query(func.count(distinct(GameSession.user_id))).join(
        User, GameSession.user_id == User.id
    ).filter(User.city == 'moscow').scalar() or 0

    region_total = User.query.filter(User.city == 'region').count()
    region_verified = User.query.filter(User.city == 'region', User.is_verified == True).count()
    region_played = db.session.query(func.count(distinct(GameSession.user_id))).join(
        User, GameSession.user_id == User.id
    ).filter(User.city == 'region').scalar() or 0

    top_moscow = User.query.filter(
        User.is_verified == True, User.total_score > 0, User.city == 'moscow'
    ).order_by(User.total_score.desc()).limit(10).all()
    top_region = User.query.filter(
        User.is_verified == True, User.total_score > 0, User.city == 'region'
    ).order_by(User.total_score.desc()).limit(10).all()

    # Full city_name breakdown
    region_breakdown = db.session.query(
        User.city_name, func.count(User.id).label('count')
    ).filter(
        User.city == 'region', User.city_name.isnot(None), User.city_name != ''
    ).group_by(User.city_name).order_by(func.count(User.id).desc()).all()

    moscow_breakdown = db.session.query(
        User.city_name, func.count(User.id).label('count')
    ).filter(
        User.city == 'moscow', User.city_name.isnot(None), User.city_name != ''
    ).group_by(User.city_name).order_by(func.count(User.id).desc()).all()

    return {
        'summary': {
            'total_users': total_users,
            'verified_users': verified_users,
            'users_played': users_played,
            'users_completed_all': users_completed_all,
            'total_levels': total_levels,
            'total_games': total_games,
            'won_games': won_games,
            'win_rate': round(won_games / total_games * 100, 1) if total_games > 0 else 0
        },
        'recent': {
            'registrations': recent_registrations,
            'games': recent_games,
            'completions': recent_completions
        },
        'charts': {
            'reg_labels': [r.date.strftime('%d.%m') if r.date else '' for r in registrations_by_day],
            'reg_data': [r.count for r in registrations_by_day],
            'games_labels': [g.date.strftime('%d.%m') if g.date else '' for g in games_by_day],
            'games_data': [g.count for g in games_by_day],
            'funnel_labels': [l['level_name'] for l in level_stats],
            'funnel_data': [l['completed'] for l in level_stats]
        },
        'level_stats': level_stats,
        'top_players': [{'rank': i+1, 'username': p.username, 'score': p.total_score} for i, p in enumerate(top_players)],
        'regions': {
            'moscow': {'total': moscow_total, 'verified': moscow_verified, 'played': moscow_played},
            'region': {'total': region_total, 'verified': region_verified, 'played': region_played},
            'top_moscow': [{'rank': i+1, 'username': p.username, 'score': p.total_score} for i, p in enumerate(top_moscow)],
            'top_region': [{'rank': i+1, 'username': p.username, 'score': p.total_score} for i, p in enumerate(top_region)],
            'moscow_breakdown': [{'name': r.city_name, 'count': r.count} for r in moscow_breakdown],
            'region_breakdown': [{'name': r.city_name, 'count': r.count} for r in region_breakdown],
        }
    }


@bp.route('/analytics/share', methods=['POST'])
@login_required
def analytics_share_settings():
    from app.models import AnalyticsShare
    import secrets

    password = request.form.get('share_password')
    if not password or len(password) < 4:
        flash('Пароль должен быть минимум 4 символа!', 'danger')
        return redirect(url_for('custom_admin.analytics'))

    share = AnalyticsShare.query.first()
    if not share:
        share = AnalyticsShare(token=secrets.token_urlsafe(16))
        db.session.add(share)

    share.set_password(password)
    share.is_active = True
    db.session.commit()

    flash('Пароль для публичной ссылки обновлён!', 'success')
    return redirect(url_for('custom_admin.analytics'))


@bp.route('/analytics/export/<format>')
@login_required
def analytics_export(format):
    from flask import Response
    from datetime import datetime
    import json
    import csv
    import io

    data = get_analytics_data()
    data['export_date'] = datetime.utcnow().isoformat()

    if format == 'json':
        return Response(
            json.dumps(data, ensure_ascii=False, indent=2),
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename=analytics_{datetime.utcnow().strftime("%Y%m%d")}.json'}
        )
    elif format == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.chart import BarChart, Reference

            wb = openpyxl.Workbook()

            # Summary sheet
            ws = wb.active
            ws.title = 'Сводка'

            # Header
            ws['A1'] = 'ROSTIC\'S Analytics'
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f'Экспорт: {datetime.utcnow().strftime("%d.%m.%Y %H:%M")}'

            # Summary data
            ws['A4'] = 'Метрика'
            ws['B4'] = 'Значение'
            ws['A4'].font = Font(bold=True)
            ws['B4'].font = Font(bold=True)

            summary_rows = [
                ('Всего пользователей', data['summary']['total_users']),
                ('Подтверждённых', data['summary']['verified_users']),
                ('Играли', data['summary']['users_played']),
                ('Прошли всё', data['summary']['users_completed_all']),
                ('Всего игр', data['summary']['total_games']),
                ('Побед', data['summary']['won_games']),
                ('Win Rate %', data['summary']['win_rate']),
            ]
            for i, (metric, value) in enumerate(summary_rows, 5):
                ws[f'A{i}'] = metric
                ws[f'B{i}'] = value

            # Levels sheet
            ws_levels = wb.create_sheet('Уровни')
            ws_levels['A1'] = 'Уровень'
            ws_levels['B1'] = 'Прошли'
            ws_levels['C1'] = 'Попыток'
            ws_levels['D1'] = 'Конверсия %'
            for cell in ['A1', 'B1', 'C1', 'D1']:
                ws_levels[cell].font = Font(bold=True)

            for i, ls in enumerate(data['level_stats'], 2):
                ws_levels[f'A{i}'] = ls['level_name']
                ws_levels[f'B{i}'] = ls['completed']
                ws_levels[f'C{i}'] = ls['attempts']
                ws_levels[f'D{i}'] = ls['conversion']

            # Top players sheet
            ws_players = wb.create_sheet('Топ игроков')
            ws_players['A1'] = '#'
            ws_players['B1'] = 'Игрок'
            ws_players['C1'] = 'Очки'
            for cell in ['A1', 'B1', 'C1']:
                ws_players[cell].font = Font(bold=True)

            for i, tp in enumerate(data['top_players'], 2):
                ws_players[f'A{i}'] = tp['rank']
                ws_players[f'B{i}'] = tp['username']
                ws_players[f'C{i}'] = tp['score']

            # Auto-width columns
            for sheet in wb.worksheets:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename=analytics_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'}
            )
        except ImportError:
            flash('Модуль openpyxl не установлен. Используйте CSV экспорт.', 'danger')
            return redirect(url_for('custom_admin.analytics'))
    else:  # CSV
        output = io.StringIO()
        writer = csv.writer(output)

        output.write('=== СВОДКА ===\n')
        writer.writerow(['Метрика', 'Значение'])
        writer.writerow(['Всего пользователей', data['summary']['total_users']])
        writer.writerow(['Подтверждённых', data['summary']['verified_users']])
        writer.writerow(['Играли', data['summary']['users_played']])
        writer.writerow(['Прошли всё', data['summary']['users_completed_all']])
        writer.writerow(['Всего игр', data['summary']['total_games']])
        writer.writerow(['Побед', data['summary']['won_games']])
        writer.writerow(['Win Rate %', data['summary']['win_rate']])

        output.write('\n=== УРОВНИ ===\n')
        writer.writerow(['Уровень', 'Прошли', 'Попыток', 'Конверсия %'])
        for ls in data['level_stats']:
            writer.writerow([ls['level_name'], ls['completed'], ls['attempts'], ls['conversion']])

        output.write('\n=== ТОП ИГРОКОВ ===\n')
        writer.writerow(['#', 'Игрок', 'Очки'])
        for tp in data['top_players']:
            writer.writerow([tp['rank'], tp['username'], tp['score']])

        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=analytics_{datetime.utcnow().strftime("%Y%m%d")}.csv'}
        )


# ============== PUBLIC ANALYTICS PAGE ==============
@bp.route('/public/<token>', methods=['GET', 'POST'])
def public_analytics(token):
    from app.models import AnalyticsShare

    share = AnalyticsShare.query.filter_by(token=token, is_active=True).first()
    if not share:
        return 'Ссылка недействительна или отключена', 404

    # Check session auth
    session_key = f'analytics_auth_{token}'
    is_authenticated = request.cookies.get(session_key) == 'true'

    if request.method == 'POST':
        password = request.form.get('password', '')
        if share.check_password(password):
            response = redirect(url_for('custom_admin.public_analytics', token=token))
            response.set_cookie(session_key, 'true', max_age=3600*24)  # 24 hours
            return response
        else:
            return render_public_login(token, error='Неверный пароль')

    if not is_authenticated:
        return render_public_login(token)

    # Show analytics
    data = get_analytics_data()
    return render_public_analytics(data, token)


def render_public_login(token, error=None):
    error_html = f'<div class="alert alert-danger">{error}</div>' if error else ''
    return f'''
    <!DOCTYPE html>
    <html lang="ru">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Аналитика ROSTIC'S</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
        <style>
            body {{ background: #1e293b; min-height: 100vh; display: flex; align-items: center; justify-content: center; }}
            .login-card {{ background: white; border-radius: 20px; padding: 40px; max-width: 400px; width: 100%; box-shadow: 0 20px 60px rgba(0,0,0,0.3); }}
            .brand {{ font-size: 1.5rem; font-weight: 700; color: #e4002b; margin-bottom: 1rem; }}
        </style>
    </head>
    <body>
        <div class="login-card">
            <div class="brand text-center">🍗 ROSTIC'S Analytics</div>
            <p class="text-center text-muted mb-4">Введите пароль для просмотра аналитики</p>
            {error_html}
            <form method="POST">
                <div class="mb-3">
                    <input type="password" name="password" class="form-control form-control-lg" placeholder="Пароль" required autofocus>
                </div>
                <button type="submit" class="btn btn-danger w-100 btn-lg">Войти</button>
            </form>
        </div>
    </body>
    </html>
    '''


def render_public_analytics(data, token):
    s = data['summary']
    r = data['recent']
    c = data['charts']
    reg = data.get('regions', {})
    m = reg.get('moscow', {})
    rg = reg.get('region', {})

    level_rows = ''
    for ls in data['level_stats']:
        conv_badge = 'success' if ls['conversion'] >= 50 else 'warning' if ls['conversion'] >= 25 else 'danger'
        level_rows += f'''
        <tr>
            <td><strong>{ls['level_name']}</strong></td>
            <td>{ls['completed']:,}</td>
            <td>{ls['attempts']:,}</td>
            <td><span class="badge bg-{conv_badge}">{ls['conversion']}%</span></td>
        </tr>
        '''

    top_rows = ''
    for tp in data['top_players'][:10]:
        medal = '🥇' if tp['rank'] == 1 else '🥈' if tp['rank'] == 2 else '🥉' if tp['rank'] == 3 else str(tp['rank'])
        top_rows += f'''
        <tr>
            <td>{medal}</td>
            <td><strong>{tp['username']}</strong></td>
            <td>{tp['score']:,}</td>
        </tr>
        '''

    top_moscow_pub_rows = ''
    for tp in reg.get('top_moscow', []):
        medal = '🥇' if tp['rank'] == 1 else '🥈' if tp['rank'] == 2 else '🥉' if tp['rank'] == 3 else str(tp['rank'])
        top_moscow_pub_rows += f'<tr><td>{medal}</td><td><strong>{tp["username"]}</strong></td><td>{tp["score"]:,}</td></tr>'

    top_region_pub_rows = ''
    for tp in reg.get('top_region', []):
        medal = '🥇' if tp['rank'] == 1 else '🥈' if tp['rank'] == 2 else '🥉' if tp['rank'] == 3 else str(tp['rank'])
        top_region_pub_rows += f'<tr><td>{medal}</td><td><strong>{tp["username"]}</strong></td><td>{tp["score"]:,}</td></tr>'

    moscow_bd_rows = ''
    for b in reg.get('moscow_breakdown', []):
        pct = round(b['count'] / m.get('total', 1) * 100, 1) if m.get('total', 0) > 0 else 0
        moscow_bd_rows += f'<tr><td><strong>{b["name"]}</strong></td><td>{b["count"]:,}</td><td>{pct}%</td></tr>'

    region_bd_rows = ''
    for idx, b in enumerate(reg.get('region_breakdown', []), 1):
        pct = round(b['count'] / rg.get('total', 1) * 100, 1) if rg.get('total', 0) > 0 else 0
        region_bd_rows += f'''<tr><td class="text-muted">{idx}</td><td><strong>{b["name"]}</strong></td><td>{b["count"]:,}</td>
            <td style="width:30%;"><div class="progress" style="height:5px;"><div class="progress-bar" style="width:{pct}%;background:#10b981;"></div></div></td><td>{pct}%</td></tr>'''

    return f'''
    <!DOCTYPE html>
    <html lang="ru">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Аналитика ROSTIC'S</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
        <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.0/font/bootstrap-icons.css" rel="stylesheet">
        <style>
            :root {{ --primary: #e4002b; }}
            body {{ background: #f8fafc; }}
            .header {{ background: linear-gradient(135deg, #e4002b 0%, #b8001f 100%); color: white; padding: 2rem; margin-bottom: 2rem; }}
            .header h1 {{ font-weight: 700; margin: 0; }}
            .stat-card {{ background: white; border-radius: 16px; padding: 1.5rem; box-shadow: 0 2px 8px rgba(0,0,0,0.05); height: 100%; }}
            .stat-value {{ font-size: 2.5rem; font-weight: 700; color: #1e293b; }}
            .stat-label {{ color: #64748b; font-size: 0.875rem; text-transform: uppercase; }}
            .card {{ border: none; border-radius: 16px; box-shadow: 0 2px 8px rgba(0,0,0,0.05); }}
            .card-header {{ background: transparent; border-bottom: 1px solid #e2e8f0; font-weight: 600; }}
            .print-only {{ display: none; }}
            @media print {{
                .no-print {{ display: none !important; }}
                .print-only {{ display: block; }}
                body {{ background: white; }}
                .header {{ background: white !important; color: #1e293b !important; border-bottom: 3px solid #e4002b; }}
            }}
        </style>
    </head>
    <body>
        <div class="header no-print">
            <div class="container">
                <div class="d-flex justify-content-between align-items-center flex-wrap gap-3">
                    <h1>🍗 ROSTIC'S Analytics</h1>
                    <button onclick="window.print()" class="btn btn-light"><i class="bi bi-file-pdf me-2"></i>Скачать PDF</button>
                </div>
            </div>
        </div>
        <div class="print-only" style="text-align: center; padding: 2rem; border-bottom: 3px solid #e4002b;">
            <h1 style="font-size: 2rem; font-weight: 700;">🍗 ROSTIC'S Analytics Report</h1>
        </div>

        <div class="container pb-5">
            <!-- Summary Cards -->
            <div class="row g-3 mb-4">
                <div class="col-md-3 col-6">
                    <div class="stat-card">
                        <div class="stat-value">{s['total_users']:,}</div>
                        <div class="stat-label">Всего пользователей</div>
                        <small class="text-muted">Подтверждённых: {s['verified_users']:,}</small>
                    </div>
                </div>
                <div class="col-md-3 col-6">
                    <div class="stat-card">
                        <div class="stat-value">{s['users_played']:,}</div>
                        <div class="stat-label">Играли</div>
                        <small class="text-muted">{round(s['users_played']/s['total_users']*100, 1) if s['total_users'] > 0 else 0}% от всех</small>
                    </div>
                </div>
                <div class="col-md-3 col-6">
                    <div class="stat-card">
                        <div class="stat-value">{s['users_completed_all']:,}</div>
                        <div class="stat-label">Прошли всё</div>
                        <small class="text-muted">Все {s['total_levels']} уровней</small>
                    </div>
                </div>
                <div class="col-md-3 col-6">
                    <div class="stat-card">
                        <div class="stat-value">{s['total_games']:,}</div>
                        <div class="stat-label">Всего игр</div>
                        <small class="text-muted">Win rate: {s['win_rate']}%</small>
                    </div>
                </div>
            </div>

            <!-- Recent Activity -->
            <div class="card mb-4">
                <div class="card-header">📊 За последние 7 дней</div>
                <div class="card-body">
                    <div class="row text-center">
                        <div class="col-4">
                            <div style="font-size: 2rem; font-weight: 700; color: #6366f1;">{r['registrations']}</div>
                            <div class="text-muted">Регистраций</div>
                        </div>
                        <div class="col-4">
                            <div style="font-size: 2rem; font-weight: 700; color: #10b981;">{r['games']}</div>
                            <div class="text-muted">Игр</div>
                        </div>
                        <div class="col-4">
                            <div style="font-size: 2rem; font-weight: 700; color: #f59e0b;">{r['completions']}</div>
                            <div class="text-muted">Уровней пройдено</div>
                        </div>
                    </div>
                </div>
            </div>

            <!-- Charts -->
            <div class="row g-3 mb-4">
                <div class="col-md-6">
                    <div class="card h-100">
                        <div class="card-header">📈 Регистрации (30 дней)</div>
                        <div class="card-body">
                            <canvas id="regChart" height="200"></canvas>
                        </div>
                    </div>
                </div>
                <div class="col-md-6">
                    <div class="card h-100">
                        <div class="card-header">🎮 Игры (30 дней)</div>
                        <div class="card-body">
                            <canvas id="gamesChart" height="200"></canvas>
                        </div>
                    </div>
                </div>
            </div>

            <!-- Funnel -->
            <div class="card mb-4">
                <div class="card-header">📊 Воронка прохождения</div>
                <div class="card-body">
                    <canvas id="funnelChart" height="100"></canvas>
                </div>
            </div>

            <!-- Tables -->
            <div class="row g-3 mb-4">
                <div class="col-md-6">
                    <div class="card h-100">
                        <div class="card-header">📋 Статистика по уровням</div>
                        <div class="table-responsive">
                            <table class="table table-sm mb-0">
                                <thead><tr><th>Уровень</th><th>Прошли</th><th>Попыток</th><th>Конверсия</th></tr></thead>
                                <tbody>{level_rows}</tbody>
                            </table>
                        </div>
                    </div>
                </div>
                <div class="col-md-6">
                    <div class="card h-100">
                        <div class="card-header">🏆 Топ-10 игроков</div>
                        <div class="table-responsive">
                            <table class="table table-sm mb-0">
                                <thead><tr><th>#</th><th>Игрок</th><th>Очки</th></tr></thead>
                                <tbody>{top_rows}</tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>

            <!-- Region Distribution -->
            <div class="card mb-4">
                <div class="card-header">📍 Распределение по регионам</div>
                <div class="card-body">
                    <div class="row">
                        <div class="col-md-4 text-center">
                            <canvas id="regionPieChart" height="250"></canvas>
                        </div>
                        <div class="col-md-8">
                            <div class="row g-3">
                                <div class="col-6">
                                    <div style="background:#f0f0ff;border-radius:12px;padding:1.25rem;">
                                        <div style="font-size:1.75rem;font-weight:700;color:#6366f1;">{m.get('total', 0):,}</div>
                                        <div class="text-muted" style="font-size:0.875rem;">Москва и МО</div>
                                        <small class="text-muted">Подтв.: {m.get('verified', 0):,} | Играли: {m.get('played', 0):,}</small>
                                    </div>
                                </div>
                                <div class="col-6">
                                    <div style="background:#f0fdf4;border-radius:12px;padding:1.25rem;">
                                        <div style="font-size:1.75rem;font-weight:700;color:#10b981;">{rg.get('total', 0):,}</div>
                                        <div class="text-muted" style="font-size:0.875rem;">Регионы</div>
                                        <small class="text-muted">Подтв.: {rg.get('verified', 0):,} | Играли: {rg.get('played', 0):,}</small>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <!-- Detailed breakdown -->
            <div class="row g-3 mb-4">
                <div class="col-md-4">
                    <div class="card h-100">
                        <div class="card-header">🏙 Москва и МО ({m.get('total', 0):,})</div>
                        <div class="table-responsive" style="max-height:500px;overflow-y:auto;">
                            <table class="table table-sm mb-0">
                                <thead><tr><th>Регион</th><th>Кол-во</th><th>%</th></tr></thead>
                                <tbody>{moscow_bd_rows if moscow_bd_rows else '<tr><td colspan="3" class="text-center text-muted">Нет данных</td></tr>'}</tbody>
                            </table>
                        </div>
                    </div>
                </div>
                <div class="col-md-8">
                    <div class="card h-100">
                        <div class="card-header">🌍 Все регионы ({rg.get('total', 0):,}) — {len(reg.get('region_breakdown', []))} регионов</div>
                        <div class="table-responsive" style="max-height:500px;overflow-y:auto;">
                            <table class="table table-sm mb-0">
                                <thead><tr><th>#</th><th>Регион</th><th>Кол-во</th><th></th><th>%</th></tr></thead>
                                <tbody>{region_bd_rows if region_bd_rows else '<tr><td colspan="5" class="text-center text-muted">Нет данных</td></tr>'}</tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>

            <!-- Top players by region -->
            <div class="row g-3">
                <div class="col-md-6">
                    <div class="card h-100">
                        <div class="card-header">🏆 Топ-10 Москва</div>
                        <div class="table-responsive">
                            <table class="table table-sm mb-0">
                                <thead><tr><th>#</th><th>Игрок</th><th>Очки</th></tr></thead>
                                <tbody>{top_moscow_pub_rows if top_moscow_pub_rows else '<tr><td colspan="3" class="text-center text-muted">Нет данных</td></tr>'}</tbody>
                            </table>
                        </div>
                    </div>
                </div>
                <div class="col-md-6">
                    <div class="card h-100">
                        <div class="card-header">🏆 Топ-10 Регионы</div>
                        <div class="table-responsive">
                            <table class="table table-sm mb-0">
                                <thead><tr><th>#</th><th>Игрок</th><th>Очки</th></tr></thead>
                                <tbody>{top_region_pub_rows if top_region_pub_rows else '<tr><td colspan="3" class="text-center text-muted">Нет данных</td></tr>'}</tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
        <script>
            new Chart(document.getElementById('regChart'), {{
                type: 'line',
                data: {{
                    labels: {c['reg_labels']},
                    datasets: [{{ label: 'Регистрации', data: {c['reg_data']}, borderColor: '#6366f1', backgroundColor: 'rgba(99, 102, 241, 0.1)', fill: true, tension: 0.4 }}]
                }},
                options: {{ responsive: true, plugins: {{ legend: {{ display: false }} }}, scales: {{ y: {{ beginAtZero: true, ticks: {{ precision: 0 }} }} }} }}
            }});

            new Chart(document.getElementById('gamesChart'), {{
                type: 'line',
                data: {{
                    labels: {c['games_labels']},
                    datasets: [{{ label: 'Игры', data: {c['games_data']}, borderColor: '#10b981', backgroundColor: 'rgba(16, 185, 129, 0.1)', fill: true, tension: 0.4 }}]
                }},
                options: {{ responsive: true, plugins: {{ legend: {{ display: false }} }}, scales: {{ y: {{ beginAtZero: true, ticks: {{ precision: 0 }} }} }} }}
            }});

            new Chart(document.getElementById('funnelChart'), {{
                type: 'bar',
                data: {{
                    labels: {c['funnel_labels']},
                    datasets: [{{ label: 'Прошли', data: {c['funnel_data']}, backgroundColor: ['rgba(99, 102, 241, 0.8)', 'rgba(16, 185, 129, 0.8)', 'rgba(245, 158, 11, 0.8)', 'rgba(239, 68, 68, 0.8)', 'rgba(6, 182, 212, 0.8)', 'rgba(168, 85, 247, 0.8)'], borderRadius: 8 }}]
                }},
                options: {{ responsive: true, plugins: {{ legend: {{ display: false }} }}, scales: {{ y: {{ beginAtZero: true, ticks: {{ precision: 0 }} }} }} }}
            }});

            new Chart(document.getElementById('regionPieChart'), {{
                type: 'doughnut',
                data: {{
                    labels: ['Москва и МО', 'Регионы'],
                    datasets: [{{
                        data: [{m.get('total', 0)}, {rg.get('total', 0)}],
                        backgroundColor: ['rgba(99, 102, 241, 0.8)', 'rgba(16, 185, 129, 0.8)'],
                        borderWidth: 0
                    }}]
                }},
                options: {{
                    responsive: true,
                    plugins: {{ legend: {{ position: 'bottom' }} }}
                }}
            }});
        </script>
    </body>
    </html>
    '''


# ============== LANDING STATS (Sakura Fest) ==============
@bp.route('/landing-stats/')
@login_required
def landing_stats():
    from sqlalchemy import func, distinct
    from datetime import datetime, timedelta
    import secrets

    # Summary stats
    total_visits = LandingVisit.query.count()
    unique_ips = db.session.query(func.count(distinct(LandingVisit.ip_address))).scalar() or 0

    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    today_visits = LandingVisit.query.filter(LandingVisit.created_at >= today_start).count()

    first_visit = db.session.query(func.min(LandingVisit.created_at)).scalar()
    if first_visit:
        days_active = max((datetime.utcnow() - first_visit).days, 1)
        avg_per_day = round(total_visits / days_active, 1)
    else:
        avg_per_day = 0

    # Visits by day (all time)
    visits_by_day = db.session.query(
        func.date(LandingVisit.created_at).label('date'),
        func.count(LandingVisit.id).label('count')
    ).group_by(
        func.date(LandingVisit.created_at)
    ).order_by(func.date(LandingVisit.created_at)).all()

    day_labels = [r.date.strftime('%d.%m') if r.date else '' for r in visits_by_day]
    day_data = [r.count for r in visits_by_day]

    # City breakdown
    city_stats = db.session.query(
        LandingVisit.city,
        func.count(LandingVisit.id).label('count')
    ).filter(LandingVisit.city.isnot(None)).group_by(
        LandingVisit.city
    ).order_by(func.count(LandingVisit.id).desc()).limit(20).all()

    from markupsafe import escape

    city_rows = ''
    for cs in city_stats:
        pct = round(cs.count / total_visits * 100, 1) if total_visits > 0 else 0
        city_rows += f'''
        <tr>
            <td><strong>{escape(cs.city)}</strong></td>
            <td>{cs.count:,}</td>
            <td><div class="progress" style="height:6px;"><div class="progress-bar" style="width:{pct}%;background:var(--primary);"></div></div></td>
            <td>{pct}%</td>
        </tr>'''

    # Recent visits (last 50)
    recent_visits = LandingVisit.query.order_by(LandingVisit.created_at.desc()).limit(50).all()
    recent_rows = ''
    for v in recent_visits:
        city_badge = f'<span class="badge badge-info">{escape(v.city)}</span>' if v.city else '<span class="text-muted">—</span>'
        fake_badge = ' <span class="badge badge-warning">seed</span>' if v.is_fake else ''
        ref = v.referrer[:40] + '...' if v.referrer and len(v.referrer) > 40 else (v.referrer or '—')
        recent_rows += f'''
        <tr>
            <td><code style="background:#f1f5f9;padding:2px 8px;border-radius:4px;">{escape(v.ip_address or "—")}</code></td>
            <td>{city_badge}{fake_badge}</td>
            <td style="max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="{escape(v.user_agent or "")}">{escape((v.user_agent or "—")[:60])}</td>
            <td>{escape(ref)}</td>
            <td>{v.created_at.strftime("%d.%m.%Y %H:%M") if v.created_at else "—"}</td>
        </tr>'''

    # Get or create share token
    share = LandingStatsShare.query.first()
    if not share:
        share = LandingStatsShare(token=secrets.token_urlsafe(16), password_hash='')
        db.session.add(share)
        db.session.commit()

    content = f'''
    <div class="page-header">
        <h1 class="page-title"><i class="bi bi-flower2 me-2"></i>Sakura Fest — Статистика лендинга</h1>
        <div class="d-flex gap-2 flex-wrap">
            <button class="btn btn-primary" data-bs-toggle="modal" data-bs-target="#shareModal"><i class="bi bi-share me-2"></i>Поделиться</button>
        </div>
    </div>

    <!-- Share Modal -->
    <div class="modal fade" id="shareModal" tabindex="-1">
        <div class="modal-dialog">
            <div class="modal-content">
                <div class="modal-header">
                    <h5 class="modal-title"><i class="bi bi-share me-2"></i>Поделиться статистикой лендинга</h5>
                    <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
                </div>
                <form method="POST" action="/admin/landing-stats/share">
                    <div class="modal-body">
                        <p class="text-muted">Создайте защищённую ссылку на статистику лендинга. Получатель сможет просмотреть графики посещений.</p>
                        <div class="mb-3">
                            <label class="form-label">Пароль для доступа</label>
                            <input type="password" name="share_password" class="form-control" placeholder="Введите пароль" required>
                        </div>
                        <div class="alert alert-info">
                            <i class="bi bi-info-circle me-2"></i>
                            <strong>Публичная ссылка:</strong><br>
                            <code id="shareUrl">{request.host_url}admin/public/landing/{share.token}</code>
                            <button type="button" class="btn btn-sm btn-outline-secondary ms-2" onclick="navigator.clipboard.writeText(document.getElementById('shareUrl').innerText)">
                                <i class="bi bi-clipboard"></i>
                            </button>
                        </div>
                    </div>
                    <div class="modal-footer">
                        <button type="button" class="btn btn-outline-secondary" data-bs-dismiss="modal">Отмена</button>
                        <button type="submit" class="btn btn-primary"><i class="bi bi-check-lg me-2"></i>Сохранить пароль</button>
                    </div>
                </form>
            </div>
        </div>
    </div>

    <!-- Summary Cards -->
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:1.5rem;margin-bottom:2rem;">
        <div class="card" style="margin:0;">
            <div class="card-body text-center">
                <div style="font-size:2rem;font-weight:700;color:var(--primary);">{total_visits:,}</div>
                <div class="text-muted" style="font-size:0.875rem;">Всего визитов</div>
            </div>
        </div>
        <div class="card" style="margin:0;">
            <div class="card-body text-center">
                <div style="font-size:2rem;font-weight:700;color:var(--success);">{unique_ips:,}</div>
                <div class="text-muted" style="font-size:0.875rem;">Уникальных IP</div>
            </div>
        </div>
        <div class="card" style="margin:0;">
            <div class="card-body text-center">
                <div style="font-size:2rem;font-weight:700;color:var(--warning);">{today_visits:,}</div>
                <div class="text-muted" style="font-size:0.875rem;">Сегодня</div>
            </div>
        </div>
        <div class="card" style="margin:0;">
            <div class="card-body text-center">
                <div style="font-size:2rem;font-weight:700;color:var(--info);">{avg_per_day}</div>
                <div class="text-muted" style="font-size:0.875rem;">В среднем / день</div>
            </div>
        </div>
    </div>

    <!-- Daily Chart -->
    <div class="card">
        <div class="card-header">
            <h3 class="card-title"><i class="bi bi-graph-up me-2"></i>Визиты по дням</h3>
        </div>
        <div class="card-body">
            <canvas id="dailyChart" height="100"></canvas>
        </div>
    </div>

    <div style="display:grid;grid-template-columns:1fr 1fr;gap:1.5rem;">
        <!-- City Breakdown -->
        <div class="card" style="margin:0;">
            <div class="card-header">
                <h3 class="card-title"><i class="bi bi-geo-alt me-2"></i>География</h3>
            </div>
            <div class="card-body" style="padding:0;">
                <table class="table">
                    <thead>
                        <tr><th>Город</th><th>Визитов</th><th></th><th>%</th></tr>
                    </thead>
                    <tbody>
                        {city_rows if city_rows else '<tr><td colspan="4" class="text-center text-muted">Нет данных о городах</td></tr>'}
                    </tbody>
                </table>
            </div>
        </div>

        <!-- Recent Visits -->
        <div class="card" style="margin:0;">
            <div class="card-header">
                <h3 class="card-title"><i class="bi bi-clock-history me-2"></i>Последние визиты</h3>
            </div>
            <div class="card-body" style="padding:0;max-height:400px;overflow-y:auto;">
                <table class="table" style="font-size:0.8rem;">
                    <thead>
                        <tr><th>IP</th><th>Город</th><th>UA</th><th>Referrer</th><th>Дата</th></tr>
                    </thead>
                    <tbody>
                        {recent_rows if recent_rows else '<tr><td colspan="5" class="text-center text-muted">Нет визитов</td></tr>'}
                    </tbody>
                </table>
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script>
        new Chart(document.getElementById('dailyChart'), {{
            type: 'bar',
            data: {{
                labels: {day_labels},
                datasets: [{{
                    label: 'Визиты',
                    data: {day_data},
                    backgroundColor: 'rgba(99, 102, 241, 0.7)',
                    borderColor: 'rgba(99, 102, 241, 1)',
                    borderWidth: 1,
                    borderRadius: 6
                }}]
            }},
            options: {{
                responsive: true,
                plugins: {{ legend: {{ display: false }} }},
                scales: {{
                    y: {{ beginAtZero: true, ticks: {{ precision: 0 }} }},
                    x: {{ grid: {{ display: false }} }}
                }}
            }}
        }});
    </script>
    '''

    return render_page('Sakura Fest — Статистика', 'landing_stats', content)


@bp.route('/landing-stats/share', methods=['POST'])
@login_required
def landing_stats_share_settings():
    import secrets

    password = request.form.get('share_password')
    if not password or len(password) < 4:
        flash('Пароль должен быть минимум 4 символа!', 'danger')
        return redirect(url_for('custom_admin.landing_stats'))

    share = LandingStatsShare.query.first()
    if not share:
        share = LandingStatsShare(token=secrets.token_urlsafe(16))
        db.session.add(share)

    share.set_password(password)
    share.is_active = True
    db.session.commit()

    flash('Пароль для публичной ссылки обновлён!', 'success')
    return redirect(url_for('custom_admin.landing_stats'))


# ============== PUBLIC LANDING STATS PAGE ==============
@bp.route('/public/landing/<token>', methods=['GET', 'POST'])
def public_landing_stats(token):
    share = LandingStatsShare.query.filter_by(token=token, is_active=True).first()
    if not share:
        return 'Ссылка недействительна или отключена', 404

    session_key = f'landing_auth_{token}'
    is_authenticated = request.cookies.get(session_key) == 'true'

    if request.method == 'POST':
        password = request.form.get('password', '')
        if share.check_password(password):
            response = redirect(url_for('custom_admin.public_landing_stats', token=token))
            response.set_cookie(session_key, 'true', max_age=3600*24)
            return response
        else:
            return render_landing_login(token, error='Неверный пароль')

    if not is_authenticated:
        return render_landing_login(token)

    return render_public_landing_page(token)


@bp.route('/public/landing/<token>/export/xlsx')
def public_landing_export_xlsx(token):
    from flask import Response
    from datetime import datetime
    import io

    share = LandingStatsShare.query.filter_by(token=token, is_active=True).first()
    if not share:
        return 'Ссылка недействительна', 404

    session_key = f'landing_auth_{token}'
    if request.cookies.get(session_key) != 'true':
        return 'Не авторизован', 401

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from sqlalchemy import func, distinct

    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = 'Сводка'

    total_visits = LandingVisit.query.count()
    unique_ips = db.session.query(func.count(distinct(LandingVisit.ip_address))).scalar() or 0
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    today_visits = LandingVisit.query.filter(LandingVisit.created_at >= today_start).count()

    header_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color='FFED1C29', end_color='FFED1C29', fill_type='solid')
    header_text = Font(bold=True, color='FFFFFF')

    ws['A1'] = 'Sakura Fest — Статистика'
    ws['A1'].font = header_font
    ws['A2'] = f'Экспорт: {datetime.utcnow().strftime("%d.%m.%Y %H:%M UTC")}'

    ws['A4'] = 'Метрика'
    ws['B4'] = 'Значение'
    ws['A4'].font = bold_font
    ws['B4'].font = bold_font

    summary_rows = [
        ('Всего визитов', total_visits),
        ('Уникальных IP', unique_ips),
        ('Сегодня', today_visits),
    ]
    for i, (metric, value) in enumerate(summary_rows, 5):
        ws[f'A{i}'] = metric
        ws[f'B{i}'] = value

    # --- Sheet 2: Daily stats ---
    ws_daily = wb.create_sheet('По дням')
    for col, title in [('A', 'Дата'), ('B', 'Визиты')]:
        cell = ws_daily[f'{col}1']
        cell.value = title
        cell.font = header_text
        cell.fill = header_fill

    visits_by_day = db.session.query(
        func.date(LandingVisit.created_at).label('date'),
        func.count(LandingVisit.id).label('count')
    ).group_by(func.date(LandingVisit.created_at)).order_by(func.date(LandingVisit.created_at)).all()

    for i, row in enumerate(visits_by_day, 2):
        ws_daily[f'A{i}'] = row.date.strftime('%d.%m.%Y') if row.date else ''
        ws_daily[f'B{i}'] = row.count

    # --- Sheet 3: Geography ---
    ws_geo = wb.create_sheet('География')
    for col, title in [('A', 'Город'), ('B', 'Визиты'), ('C', '%')]:
        cell = ws_geo[f'{col}1']
        cell.value = title
        cell.font = header_text
        cell.fill = header_fill

    city_stats = db.session.query(
        LandingVisit.city,
        func.count(LandingVisit.id).label('count')
    ).filter(LandingVisit.city.isnot(None), LandingVisit.city != '').group_by(
        LandingVisit.city
    ).order_by(func.count(LandingVisit.id).desc()).all()

    for i, cs in enumerate(city_stats, 2):
        pct = round(cs.count / total_visits * 100, 1) if total_visits > 0 else 0
        ws_geo[f'A{i}'] = cs.city
        ws_geo[f'B{i}'] = cs.count
        ws_geo[f'C{i}'] = pct

    # --- Sheet 4: All visitors ---
    ws_visitors = wb.create_sheet('Посетители')
    visitor_headers = ['#', 'IP-адрес', 'Город', 'Регион', 'Страна', 'User-Agent', 'Дата']
    for col_idx, title in enumerate(visitor_headers, 1):
        cell = ws_visitors.cell(row=1, column=col_idx)
        cell.value = title
        cell.font = header_text
        cell.fill = header_fill

    all_visits = LandingVisit.query.order_by(LandingVisit.created_at.desc()).all()
    for i, v in enumerate(all_visits, 2):
        ws_visitors.cell(row=i, column=1, value=i - 1)
        ws_visitors.cell(row=i, column=2, value=v.ip_address or '')
        ws_visitors.cell(row=i, column=3, value=v.city or '')
        ws_visitors.cell(row=i, column=4, value=v.region or '')
        ws_visitors.cell(row=i, column=5, value=v.country or '')
        ws_visitors.cell(row=i, column=6, value=v.user_agent or '')
        ws_visitors.cell(row=i, column=7, value=v.created_at.strftime('%d.%m.%Y %H:%M') if v.created_at else '')

    # Auto-width columns
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[col_letter].width = min(max_length + 3, 60)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f'sakura_fest_stats_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={fname}'}
    )


def render_landing_login(token, error=None):
    error_html = f'<div style="background:rgba(237,28,41,0.08);color:#C41420;border:1px solid rgba(237,28,41,0.2);border-radius:12px;padding:12px 16px;margin-bottom:16px;font-size:14px;text-align:center;">{error}</div>' if error else ''
    return f'''
    <!DOCTYPE html>
    <html lang="ru">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Sakura Fest — Статистика</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.0/font/bootstrap-icons.css" rel="stylesheet">
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
            *, *::before, *::after {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{
                font-family: 'Inter', -apple-system, sans-serif;
                background: #FFF0F3;
                min-height: 100vh;
                display: flex;
                align-items: center;
                justify-content: center;
                position: relative;
                overflow: hidden;
            }}
            body::before {{
                content: '';
                position: fixed;
                inset: 0;
                background:
                    radial-gradient(ellipse at 30% 20%, rgba(244,166,184,0.4) 0%, transparent 50%),
                    radial-gradient(ellipse at 70% 80%, rgba(255,200,215,0.3) 0%, transparent 50%);
                z-index: 0;
            }}
            .petal {{
                position: fixed;
                width: 14px; height: 14px;
                background: radial-gradient(ellipse at 30% 30%, #FFB7D5, #E8819E);
                border-radius: 50% 0 50% 50%;
                opacity: 0;
                pointer-events: none;
                z-index: 0;
                animation: pd linear infinite;
            }}
            @keyframes pd {{
                0%   {{ transform: translate(0, -60px) rotate(0deg); opacity: 0; }}
                8%   {{ opacity: 0.6; }}
                50%  {{ transform: translate(var(--sway), 50vh) rotate(180deg); opacity: 0.4; }}
                100% {{ transform: translate(calc(var(--sway) * -0.5), calc(100vh + 60px)) rotate(360deg); opacity: 0; }}
            }}
            .login-card {{
                position: relative;
                z-index: 1;
                background: linear-gradient(180deg, rgba(255,255,255,0.85) 0%, rgba(255,240,243,0.7) 100%);
                backdrop-filter: blur(24px);
                -webkit-backdrop-filter: blur(24px);
                border: 1px solid rgba(255,255,255,0.7);
                border-radius: 28px;
                padding: 48px 36px 40px;
                max-width: 400px;
                width: calc(100% - 40px);
                box-shadow: 0 24px 80px rgba(180,100,120,0.15);
                text-align: center;
            }}
            .login-card::before {{
                content: '';
                position: absolute;
                top: -1px; left: 20%; right: 20%;
                height: 2px;
                background: linear-gradient(90deg, transparent, #ED1C29, transparent);
                border-radius: 1px;
            }}
            .brand-icon {{ font-size: 40px; margin-bottom: 12px; display: block; }}
            .brand-title {{ font-size: 22px; font-weight: 800; color: #ED1C29; margin-bottom: 4px; }}
            .brand-sub {{ font-size: 13px; color: #94a3b8; margin-bottom: 28px; }}
            input[type="password"] {{
                width: 100%;
                padding: 14px 18px;
                border: 1px solid rgba(237,28,41,0.15);
                border-radius: 14px;
                font-size: 16px;
                font-family: inherit;
                background: rgba(255,255,255,0.7);
                outline: none;
                transition: border-color 0.2s, box-shadow 0.2s;
                margin-bottom: 16px;
            }}
            input[type="password"]:focus {{
                border-color: #ED1C29;
                box-shadow: 0 0 0 4px rgba(237,28,41,0.1);
            }}
            .submit-btn {{
                width: 100%;
                padding: 14px;
                background: #ED1C29;
                color: white;
                border: none;
                border-radius: 14px;
                font-size: 16px;
                font-weight: 700;
                cursor: pointer;
                transition: all 0.2s;
                box-shadow: 0 4px 16px rgba(237,28,41,0.3);
            }}
            .submit-btn:hover {{
                transform: translateY(-1px);
                box-shadow: 0 6px 24px rgba(237,28,41,0.4);
            }}
        </style>
    </head>
    <body>
        <div class="petal" style="left:10%;--sway:40px;animation-duration:12s;animation-delay:0s;"></div>
        <div class="petal" style="left:25%;--sway:-35px;animation-duration:15s;animation-delay:2s;width:9px;height:9px;"></div>
        <div class="petal" style="left:55%;--sway:50px;animation-duration:13s;animation-delay:1s;width:20px;height:20px;"></div>
        <div class="petal" style="left:75%;--sway:-30px;animation-duration:14s;animation-delay:3s;"></div>
        <div class="petal" style="left:90%;--sway:25px;animation-duration:16s;animation-delay:5s;width:9px;height:9px;"></div>

        <div class="login-card">
            <span class="brand-icon">🌸</span>
            <div class="brand-title">Sakura Fest</div>
            <div class="brand-sub">Введите пароль для просмотра статистики</div>
            {error_html}
            <form method="POST">
                <input type="password" name="password" placeholder="Пароль" required autofocus>
                <button type="submit" class="submit-btn">Войти</button>
            </form>
        </div>
    </body>
    </html>
    '''


def render_public_landing_page(token):
    from sqlalchemy import func, distinct
    from datetime import datetime
    from markupsafe import escape

    total_visits = LandingVisit.query.count()
    unique_ips = db.session.query(func.count(distinct(LandingVisit.ip_address))).scalar() or 0

    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    today_visits = LandingVisit.query.filter(LandingVisit.created_at >= today_start).count()

    first_visit = db.session.query(func.min(LandingVisit.created_at)).scalar()
    if first_visit:
        days_active = max((datetime.utcnow() - first_visit).days, 1)
        avg_per_day = round(total_visits / days_active, 1)
    else:
        avg_per_day = 0

    visits_by_day = db.session.query(
        func.date(LandingVisit.created_at).label('date'),
        func.count(LandingVisit.id).label('count')
    ).group_by(func.date(LandingVisit.created_at)).order_by(func.date(LandingVisit.created_at)).all()

    day_labels = [r.date.strftime('%d.%m') if r.date else '' for r in visits_by_day]
    day_data = [r.count for r in visits_by_day]

    city_stats = db.session.query(
        LandingVisit.city,
        func.count(LandingVisit.id).label('count')
    ).filter(LandingVisit.city.isnot(None)).group_by(LandingVisit.city).order_by(
        func.count(LandingVisit.id).desc()
    ).limit(20).all()

    city_rows = ''
    for cs in city_stats:
        pct = round(cs.count / total_visits * 100, 1) if total_visits > 0 else 0
        city_rows += f'''<tr>
            <td><strong>{escape(cs.city)}</strong></td>
            <td>{cs.count:,}</td>
            <td><div class="pct-bar"><div class="pct-fill" style="width:{pct}%"></div></div></td>
            <td>{pct}%</td>
        </tr>'''

    # Detailed visitors table (all visits, no referrer, no seed badge)
    all_visits = LandingVisit.query.order_by(LandingVisit.created_at.desc()).all()
    visitor_rows = ''
    for i, v in enumerate(all_visits):
        city_text = escape(v.city) if v.city else '<span class="dim">—</span>'
        country_text = escape(v.country) if v.country else '<span class="dim">—</span>'
        region_text = escape(v.region) if v.region else '<span class="dim">—</span>'
        ua_short = escape((v.user_agent or '—')[:80])
        ua_full = escape(v.user_agent or '')
        date_str = v.created_at.strftime('%d.%m.%Y %H:%M') if v.created_at else '—'
        visitor_rows += f'''<tr>
            <td class="row-num">{i + 1}</td>
            <td><code class="ip-code">{escape(v.ip_address or "—")}</code></td>
            <td>{city_text}</td>
            <td>{region_text}</td>
            <td>{country_text}</td>
            <td class="ua-cell" title="{ua_full}">{ua_short}</td>
            <td class="date-cell">{date_str}</td>
        </tr>'''

    return f'''
    <!DOCTYPE html>
    <html lang="ru">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Sakura Fest — Статистика</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.0/font/bootstrap-icons.css" rel="stylesheet">
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');

            :root {{
                --red: #ED1C29;
                --red-dark: #C41420;
                --pink-light: #FFF0F3;
                --pink-mid: #FADADD;
                --pink-deep: #F4A6B8;
                --glass-bg: rgba(255,205,220,0.35);
                --glass-border: rgba(255,255,255,0.6);
                --card-shadow: 0 8px 32px rgba(180,100,120,0.12);
            }}

            *, *::before, *::after {{ margin: 0; padding: 0; box-sizing: border-box; }}

            body {{
                font-family: 'Inter', -apple-system, sans-serif;
                background: var(--pink-light);
                color: #1e293b;
                min-height: 100vh;
                position: relative;
            }}

            /* Sakura background */
            body::before {{
                content: '';
                position: fixed;
                inset: 0;
                background:
                    radial-gradient(ellipse at 20% 20%, rgba(244,166,184,0.3) 0%, transparent 50%),
                    radial-gradient(ellipse at 80% 80%, rgba(255,200,215,0.25) 0%, transparent 50%),
                    linear-gradient(180deg, var(--pink-light) 0%, #fff5f7 40%, #fef2f4 100%);
                z-index: -1;
            }}

            /* Floating petals */
            .petal {{
                position: fixed;
                width: 14px; height: 14px;
                background: radial-gradient(ellipse at 30% 30%, #FFB7D5, #E8819E);
                border-radius: 50% 0 50% 50%;
                opacity: 0;
                pointer-events: none;
                z-index: 0;
                animation: petalDrift linear infinite;
            }}
            @keyframes petalDrift {{
                0%   {{ transform: translate(0, -60px) rotate(0deg); opacity: 0; }}
                8%   {{ opacity: 0.6; }}
                50%  {{ transform: translate(var(--sway), 50vh) rotate(180deg); opacity: 0.4; }}
                100% {{ transform: translate(calc(var(--sway) * -0.5), calc(100vh + 60px)) rotate(360deg); opacity: 0; }}
            }}

            .container {{
                max-width: 1100px;
                margin: 0 auto;
                padding: 0 20px;
                position: relative;
                z-index: 1;
            }}

            /* Header */
            .header {{
                text-align: center;
                padding: 48px 20px 32px;
            }}
            .header-brand {{
                display: inline-flex;
                align-items: center;
                gap: 12px;
                margin-bottom: 12px;
            }}
            .header-brand img {{
                height: 48px;
                filter: drop-shadow(0 0 12px rgba(237,28,41,0.5));
            }}
            .header h1 {{
                font-size: 28px;
                font-weight: 800;
                color: var(--red);
                letter-spacing: -0.5px;
            }}
            .header p {{
                color: #94a3b8;
                font-size: 14px;
                margin-top: 4px;
            }}
            .header-actions {{
                display: flex;
                gap: 10px;
                justify-content: center;
                margin-top: 16px;
            }}
            .export-btn {{
                display: inline-flex;
                align-items: center;
                gap: 6px;
                padding: 9px 18px;
                border-radius: 12px;
                font-size: 13px;
                font-weight: 600;
                font-family: inherit;
                cursor: pointer;
                transition: all 0.2s;
                border: 1px solid rgba(237,28,41,0.2);
                background: rgba(255,255,255,0.7);
                color: var(--red);
                text-decoration: none;
                backdrop-filter: blur(8px);
            }}
            .export-btn:hover {{
                background: var(--red);
                color: white;
                border-color: var(--red);
                transform: translateY(-1px);
                box-shadow: 0 4px 12px rgba(237,28,41,0.3);
            }}

            /* Glass card */
            .glass {{
                background: linear-gradient(180deg, rgba(255,255,255,0.75) 0%, rgba(255,240,243,0.6) 100%);
                backdrop-filter: blur(24px);
                -webkit-backdrop-filter: blur(24px);
                border: 1px solid var(--glass-border);
                border-radius: 24px;
                box-shadow: var(--card-shadow);
                overflow: hidden;
                margin-bottom: 20px;
            }}
            .glass-header {{
                padding: 16px 24px;
                border-bottom: 1px solid rgba(237,28,41,0.1);
                font-weight: 700;
                font-size: 15px;
                color: #334155;
                display: flex;
                align-items: center;
                gap: 8px;
            }}
            .glass-header i {{ color: var(--red); }}
            .glass-body {{ padding: 20px 24px; }}
            .glass-body-flush {{ padding: 0; }}

            /* Stat cards */
            .stats-grid {{
                display: grid;
                grid-template-columns: repeat(4, 1fr);
                gap: 16px;
                margin-bottom: 20px;
            }}
            .stat-card {{
                background: linear-gradient(180deg, rgba(255,255,255,0.8) 0%, rgba(255,240,243,0.5) 100%);
                backdrop-filter: blur(16px);
                border: 1px solid var(--glass-border);
                border-radius: 20px;
                padding: 24px 16px;
                text-align: center;
                box-shadow: 0 4px 16px rgba(180,100,120,0.08);
                transition: transform 0.2s;
            }}
            .stat-card:hover {{ transform: translateY(-2px); }}
            .stat-value {{
                font-size: 36px;
                font-weight: 800;
                color: var(--red);
                line-height: 1;
                margin-bottom: 6px;
            }}
            .stat-label {{
                font-size: 12px;
                font-weight: 600;
                color: #94a3b8;
                text-transform: uppercase;
                letter-spacing: 1px;
            }}

            /* Tables */
            .data-table {{
                width: 100%;
                border-collapse: collapse;
                font-size: 13px;
            }}
            .data-table thead th {{
                background: rgba(237,28,41,0.05);
                padding: 10px 14px;
                font-weight: 700;
                font-size: 11px;
                color: #64748b;
                text-transform: uppercase;
                letter-spacing: 0.5px;
                border-bottom: 2px solid rgba(237,28,41,0.1);
                text-align: left;
                position: sticky;
                top: 0;
                z-index: 2;
            }}
            .data-table tbody td {{
                padding: 8px 14px;
                border-bottom: 1px solid rgba(0,0,0,0.04);
                vertical-align: middle;
            }}
            .data-table tbody tr:hover {{ background: rgba(237,28,41,0.03); }}
            .row-num {{ color: #cbd5e1; font-weight: 600; font-size: 11px; width: 40px; }}
            .ip-code {{
                background: rgba(237,28,41,0.06);
                color: var(--red-dark);
                padding: 2px 8px;
                border-radius: 6px;
                font-size: 12px;
                font-family: 'SF Mono', 'Fira Code', monospace;
            }}
            .ua-cell {{
                max-width: 250px;
                overflow: hidden;
                text-overflow: ellipsis;
                white-space: nowrap;
                color: #94a3b8;
                font-size: 11px;
            }}
            .date-cell {{ white-space: nowrap; color: #64748b; font-size: 12px; }}
            .dim {{ color: #cbd5e1; }}

            /* Percent bar */
            .pct-bar {{
                height: 6px;
                background: rgba(237,28,41,0.08);
                border-radius: 3px;
                overflow: hidden;
                min-width: 80px;
            }}
            .pct-fill {{
                height: 100%;
                background: linear-gradient(90deg, var(--red), var(--pink-deep));
                border-radius: 3px;
                transition: width 0.6s ease;
            }}

            /* Scrollable visitors */
            .visitors-scroll {{
                max-height: 600px;
                overflow-y: auto;
            }}
            .visitors-scroll::-webkit-scrollbar {{ width: 6px; }}
            .visitors-scroll::-webkit-scrollbar-track {{ background: transparent; }}
            .visitors-scroll::-webkit-scrollbar-thumb {{ background: var(--pink-mid); border-radius: 3px; }}

            @media (max-width: 768px) {{
                .stats-grid {{ grid-template-columns: repeat(2, 1fr); }}
                .stat-value {{ font-size: 28px; }}
                .header h1 {{ font-size: 22px; }}
                .header-actions {{ flex-wrap: wrap; }}
            }}

            @media print {{
                .no-print {{ display: none !important; }}
                body, body::before {{ background: white !important; }}
                .glass {{ box-shadow: none; border: 1px solid #e2e8f0; }}
                .petal {{ display: none; }}
            }}
        </style>
    </head>
    <body>
        <!-- Sakura petals -->
        <div class="petal" style="left:5%;--sway:40px;animation-duration:12s;animation-delay:0s;"></div>
        <div class="petal" style="left:15%;--sway:-30px;animation-duration:15s;animation-delay:2s;width:9px;height:9px;"></div>
        <div class="petal" style="left:30%;--sway:50px;animation-duration:13s;animation-delay:4s;"></div>
        <div class="petal" style="left:50%;--sway:-45px;animation-duration:14s;animation-delay:1s;width:20px;height:20px;"></div>
        <div class="petal" style="left:65%;--sway:35px;animation-duration:16s;animation-delay:3s;"></div>
        <div class="petal" style="left:80%;--sway:-25px;animation-duration:11s;animation-delay:5s;width:9px;height:9px;"></div>
        <div class="petal" style="left:92%;--sway:30px;animation-duration:14s;animation-delay:7s;"></div>
        <div class="petal" style="left:40%;--sway:-40px;animation-duration:17s;animation-delay:6s;width:20px;height:20px;"></div>

        <div class="container">
            <!-- Header -->
            <div class="header">
                <h1><i class="bi bi-flower2"></i> Sakura Fest</h1>
                <p>Статистика посещений лендинга</p>
                <div class="header-actions no-print">
                    <button onclick="downloadPDF()" id="pdfBtn" class="export-btn">
                        <i class="bi bi-file-pdf"></i> PDF
                    </button>
                    <a href="/admin/public/landing/{token}/export/xlsx" class="export-btn">
                        <i class="bi bi-file-earmark-excel"></i> Excel
                    </a>
                </div>
            </div>

            <!-- Stats -->
            <div class="stats-grid">
                <div class="stat-card">
                    <div class="stat-value">{total_visits:,}</div>
                    <div class="stat-label">Всего визитов</div>
                </div>
                <div class="stat-card">
                    <div class="stat-value">{unique_ips:,}</div>
                    <div class="stat-label">Уникальных IP</div>
                </div>
                <div class="stat-card">
                    <div class="stat-value">{today_visits:,}</div>
                    <div class="stat-label">Сегодня</div>
                </div>
                <div class="stat-card">
                    <div class="stat-value">{avg_per_day}</div>
                    <div class="stat-label">В среднем / день</div>
                </div>
            </div>

            <!-- Chart -->
            <div class="glass">
                <div class="glass-header"><i class="bi bi-graph-up"></i> Визиты по дням</div>
                <div class="glass-body">
                    <canvas id="dailyChart" height="100"></canvas>
                </div>
            </div>

            <!-- Geography -->
            <div class="glass">
                <div class="glass-header"><i class="bi bi-geo-alt"></i> География</div>
                <div class="glass-body-flush">
                    <table class="data-table">
                        <thead><tr><th>Город</th><th>Визитов</th><th></th><th>%</th></tr></thead>
                        <tbody>{city_rows if city_rows else '<tr><td colspan="4" style="text-align:center;color:#94a3b8;padding:24px;">Нет данных о городах</td></tr>'}</tbody>
                    </table>
                </div>
            </div>

            <!-- Detailed visitors table -->
            <div class="glass">
                <div class="glass-header"><i class="bi bi-people"></i> Все посетители ({total_visits:,})</div>
                <div class="glass-body-flush visitors-scroll">
                    <table class="data-table">
                        <thead>
                            <tr>
                                <th>#</th>
                                <th>IP-адрес</th>
                                <th>Город</th>
                                <th>Регион</th>
                                <th>Страна</th>
                                <th>Устройство</th>
                                <th>Дата</th>
                            </tr>
                        </thead>
                        <tbody>
                            {visitor_rows if visitor_rows else '<tr><td colspan="7" style="text-align:center;color:#94a3b8;padding:24px;">Нет посетителей</td></tr>'}
                        </tbody>
                    </table>
                </div>
            </div>

            <div style="text-align:center;padding:32px 0;color:#cbd5e1;font-size:12px;">
                Sakura Fest — ROSTIC'S &copy; 2026
            </div>
        </div>

        <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
        <script src="https://cdnjs.cloudflare.com/ajax/libs/html2pdf.js/0.10.1/html2pdf.bundle.min.js"></script>
        <script>
            new Chart(document.getElementById('dailyChart'), {{
                type: 'bar',
                data: {{
                    labels: {day_labels},
                    datasets: [{{
                        label: 'Визиты',
                        data: {day_data},
                        backgroundColor: 'rgba(237, 28, 41, 0.65)',
                        borderColor: 'rgba(237, 28, 41, 1)',
                        borderWidth: 1,
                        borderRadius: 8,
                        hoverBackgroundColor: 'rgba(237, 28, 41, 0.85)'
                    }}]
                }},
                options: {{
                    responsive: true,
                    plugins: {{
                        legend: {{ display: false }},
                        tooltip: {{
                            backgroundColor: 'rgba(30,30,30,0.9)',
                            titleFont: {{ weight: '600' }},
                            cornerRadius: 8,
                            padding: 12
                        }}
                    }},
                    scales: {{
                        y: {{
                            beginAtZero: true,
                            ticks: {{ precision: 0, color: '#94a3b8' }},
                            grid: {{ color: 'rgba(237,28,41,0.06)' }}
                        }},
                        x: {{
                            grid: {{ display: false }},
                            ticks: {{ color: '#94a3b8' }}
                        }}
                    }}
                }}
            }});

            function downloadPDF() {{
                var btn = document.getElementById('pdfBtn');
                btn.textContent = 'Генерация PDF...';
                btn.disabled = true;

                var element = document.querySelector('.container');
                var opt = {{
                    margin:       [10, 10, 10, 10],
                    filename:     'sakura-fest-stats.pdf',
                    image:        {{ type: 'jpeg', quality: 0.95 }},
                    html2canvas:  {{ scale: 2, useCORS: true, backgroundColor: '#FFF0F3' }},
                    jsPDF:        {{ unit: 'mm', format: 'a4', orientation: 'portrait' }},
                    pagebreak:    {{ mode: ['avoid-all', 'css', 'legacy'] }}
                }};

                html2pdf().set(opt).from(element).save().then(function() {{
                    btn.innerHTML = '<i class="bi bi-file-pdf"></i> Скачать PDF';
                    btn.disabled = false;
                }}).catch(function() {{
                    btn.innerHTML = '<i class="bi bi-file-pdf"></i> Скачать PDF';
                    btn.disabled = false;
                }});
            }}
        </script>
    </body>
    </html>
    '''
